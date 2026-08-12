"""Fournisseurs page for local AR / CHR supplier tracking."""

from __future__ import annotations

import hashlib
import html
from io import BytesIO

import streamlit as st
from openpyxl import Workbook, load_workbook

from core.suppliers import (
    CARD_STATES,
    DUPLICATE_STATES,
    MEDIA_STATES,
    PAYPAL_STATES,
    BIG_QUANTITY_STATES,
    CONVERSATION_TAGS,
    MESSAGE_AUTHORS,
    NEGOTIATION_ACTORS,
    NEGOTIATION_EVENT_TYPES,
    SUPPLIER_CURRENCIES,
    SUPPLIER_SPECIALTIES,
    SUPPLIER_STATUSES,
    SUPPLIER_TYPES,
    TRACKING_STATES,
    _parse_best_for_tags,
    _parse_first_order_quantity,
    _parse_iso_date,
    _parse_market_position_rank,
    calculate_offer,
    confidence_score,
    default_supplier,
    extract_conversation_from_text,
    extract_offer_from_text,
    make_price_history_entry,
    negotiation_chart_series,
    negotiation_stats,
    now_iso,
    normalize_supplier,
    normalize_supplier_identity,
    normalize_negotiation_event,
    normalize_conversation_message,
    pokestock_score,
    supplier_name_matches,
    smart_rankings,
    supplier_rankings,
)
from services.suppliers_data import (
    apply_supplier_import,
    delete_supplier,
    delete_conversation_message,
    delete_negotiation_event,
    delete_pending_import,
    detect_duplicate,
    duplicate_offer,
    load_suppliers,
    save_conversation_message,
    save_negotiation_event,
    save_suppliers,
    repair_existing_supplier_reviews,
    duplicate_supplier_groups,
    ignore_duplicate_pair,
    merge_suppliers,
    upsert_supplier,
)


def _display_cell(value):
    if value is None:
        return "—"
    try:
        if value != value:
            return "—"
    except Exception:
        pass
    if isinstance(value, float):
        return f"{value:.2f}".rstrip("0").rstrip(".")
    return str(value)


def _prepare_supplier_display_rows(rows):
    prepared = []
    for row in rows or []:
        prepared.append({str(key): _display_cell(value) for key, value in dict(row).items()})
    return prepared


def _money(value, currency="EUR"):
    if value is None:
        return "—"
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return "—"
    suffix = "€" if currency == "EUR" else currency
    return f"{amount:.2f} {suffix}"


def _money_eur(value):
    if value is None:
        return "À confirmer"
    try:
        return f"{float(value):.2f} €"
    except (TypeError, ValueError):
        return "À confirmer"


def _eur_source_label(eur_value, source_value=None, source_currency="", missing="À confirmer"):
    if eur_value is None:
        return missing
    label = _money_eur(eur_value)
    if source_value is not None and source_currency and source_currency != "EUR":
        label += f" · {_format_money_value(source_value, source_currency, 0)}"
    return label


def _scenario_resale_total(supplier, settings):
    quantity = supplier.get("quantity_reference") or 200
    scenario = (settings or {}).get("resale_scenario", "reference")
    per_card = {"prudent": 2.5, "reference": 3.0, "optimiste": 3.5}.get(
        scenario,
        float((settings or {}).get("custom_resale_per_card_eur") or 3.0),
    )
    if scenario == "reference" and normalize_supplier_identity(supplier.get("nom", "")) == "kkexportjapan":
        return 655.0
    return float(quantity or 200) * per_card


def _scenario_metrics(supplier, settings):
    landed = supplier.get("landed_cost_estimated_eur")
    if landed is None and supplier.get("landed_cost_estimated_min_eur") is not None and supplier.get("landed_cost_estimated_max_eur") is not None:
        landed = (supplier.get("landed_cost_estimated_min_eur") + supplier.get("landed_cost_estimated_max_eur")) / 2
    resale = _scenario_resale_total(supplier, settings)
    quantity = supplier.get("quantity_reference") or 200
    if landed is None or not resale:
        return {"resale": resale, "margin": None, "purchase_pct": None, "margin_per_card": None}
    margin = resale - landed
    return {
        "resale": resale,
        "margin": margin,
        "purchase_pct": landed / resale * 100,
        "margin_per_card": margin / float(quantity or 200),
    }


def _num(value, default=0.0):
    try:
        return float(value if value is not None else default)
    except (TypeError, ValueError):
        return float(default)


def _supplier_widget_key(prefix, group_index, identity, keep_id="", merge_id=""):
    raw = f"{prefix}|{group_index}|{identity}|{keep_id}|{merge_id}"
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]
    return f"{prefix}_{digest}"


def _badge(label, tone="neutral"):
    return f'<span class="ps-status-badge ps-status-{tone}">{html.escape(str(label))}</span>'


def _clean_text(value, empty="—"):
    if value in (None, "", [], "None", "null"):
        return empty
    text = str(value)
    mapping = {
        "low": "Faible",
        "medium": "Moyen",
        "high": "Élevé",
        "unknown": "Non renseigné",
        "yes": "Oui",
        "true": "Oui",
        "no": "Non",
        "false": "Non",
        "oui": "Oui",
        "non": "Non",
        "inconnu": "Non renseigné",
    }
    return mapping.get(text.strip().lower(), text)


def _risk_label(value):
    return _clean_text(value or "unknown")


def _currency_symbol(currency):
    return {"EUR": "€", "JPY": "¥", "USD": "$"}.get(currency or "EUR", currency or "")


def _format_unit_price(value, currency):
    if value is None:
        return None
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    symbol = _currency_symbol(currency)
    if currency == "JPY":
        return f"{symbol}{amount:.0f} / carte"
    if currency == "USD":
        return f"{symbol}{amount:.2f} / carte"
    if currency == "EUR":
        return f"{amount:.2f} {symbol} / carte"
    return f"{amount:.2f} {symbol} / carte".strip()


def _format_money_value(value, currency="EUR", decimals=2):
    if value is None:
        return "—"
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return "—"
    symbol = _currency_symbol(currency)
    if currency == "JPY":
        return f"{symbol}{amount:.0f}"
    if currency == "USD":
        return f"{symbol}{amount:.2f}"
    if currency == "EUR":
        return f"{amount:.{decimals}f} {symbol}"
    return f"{amount:.{decimals}f} {symbol}".strip()


def _format_eur_range(min_value, max_value, average=None, suffix=""):
    if min_value is not None and max_value is not None:
        return f"{float(min_value):.2f}–{float(max_value):.2f} €{suffix}"
    if average is not None:
        return f"{float(average):.2f} €{suffix}"
    return "Incomplet"


def _format_percent_range(min_value, max_value, average=None):
    if min_value is not None and max_value is not None:
        return f"{float(min_value):.1f}–{float(max_value):.1f} %"
    if average is not None:
        return f"{float(average):.1f} %"
    return "—"


def _landed_cost_label(supplier):
    return _format_eur_range(
        supplier.get("landed_cost_per_card_min_eur"),
        supplier.get("landed_cost_per_card_max_eur"),
        supplier.get("landed_cost_per_card_eur"),
        " / carte",
    )


def _landed_total_label(supplier):
    return _format_eur_range(
        supplier.get("landed_cost_estimated_min_eur"),
        supplier.get("landed_cost_estimated_max_eur"),
        supplier.get("landed_cost_estimated_eur"),
    )


def _cards_price_label(supplier):
    value = supplier.get("cards_price_source")
    currency = supplier.get("cards_price_source_currency") or supplier.get("devise") or "EUR"
    qty = supplier.get("quantity_reference") or 200
    if value is None:
        return "Prix cartes inconnu"
    unit = float(value) / float(qty) if qty else None
    return f"{_format_money_value(value, currency, 0)} · {_format_unit_price(unit, currency) if unit is not None else '—'}"


def _shipping_label(supplier):
    if supplier.get("shipping_price_source") is not None:
        return _format_money_value(supplier.get("shipping_price_source"), supplier.get("shipping_price_source_currency") or supplier.get("devise") or "EUR", 0)
    if supplier.get("shipping_status") == "about_20_usd":
        return "env. $20"
    if supplier.get("shipping_status") == "included":
        return "Incluse"
    if supplier.get("shipping_status") == "unknown":
        return "Inconnue"
    return "—"


def _paypal_label(supplier):
    if supplier.get("paypal_fee_percent") is not None:
        return f"{float(supplier.get('paypal_fee_percent')):.1f} %"
    if supplier.get("paypal_fee_source") is not None:
        return _format_money_value(supplier.get("paypal_fee_source"), supplier.get("paypal_fee_source_currency") or supplier.get("devise") or "EUR")
    if supplier.get("paypal_gs") == "inconnu":
        return "Inconnu"
    return "—"


def _supplier_total_label(supplier):
    if supplier.get("supplier_total_source") is None:
        return "Incomplet"
    return _format_money_value(supplier.get("supplier_total_source"), supplier.get("supplier_total_source_currency") or supplier.get("devise") or "EUR", 0)


def _margin_label(supplier):
    if supplier.get("estimated_margin_min_eur") is not None and supplier.get("estimated_margin_max_eur") is not None:
        return f"{supplier.get('estimated_margin_min_eur'):.0f}–{supplier.get('estimated_margin_max_eur'):.0f} €"
    if supplier.get("estimated_margin_eur") is not None:
        return f"{supplier.get('estimated_margin_eur'):.0f} €"
    return "—"


def _supplier_unit_price_label(supplier):
    eur_label = _format_unit_price(supplier.get("prix_unitaire_eur"), "EUR")
    if eur_label:
        return eur_label
    raw_label = _format_unit_price(supplier.get("prix_unitaire_estime"), supplier.get("devise") or "EUR")
    if raw_label:
        return raw_label
    status = supplier.get("price_status")
    if status == "devis_en_attente":
        return "Devis en attente"
    if status == "prix_a_confirmer":
        return "Prix à confirmer"
    if status == "non_communique":
        return "Prix non communiqué"
    return "Prix à renseigner"


def _supplier_price_hint(supplier):
    offer = calculate_offer(supplier, 200)
    if offer.get("available") and offer.get("source_label"):
        suffix = offer.get("currency") or supplier.get("devise") or ""
        return f"{offer.get('source_label')} · {suffix}".strip()
    if supplier.get("prix_unitaire_eur") is not None:
        return "Comparable en EUR"
    if supplier.get("prix_unitaire_estime") is not None and supplier.get("devise") != "EUR":
        return "Conversion EUR manquante"
    if supplier.get("price_status") == "devis_en_attente":
        return "Prix attendu du fournisseur"
    if supplier.get("price_status") == "prix_a_confirmer":
        return "Montant à confirmer"
    if supplier.get("price_status") == "non_communique":
        return "Prix non communiqué"
    return "Calcul indisponible"


def _priority_score(supplier):
    status = supplier.get("statut")
    risk = supplier.get("risk_level") or "unknown"
    score = {
        "top fournisseur": 35,
        "testé": 22,
        "actif": 15,
        "en attente": 8,
        "à contacter": 5,
        "à éviter": -100,
        "archivé": -100,
    }.get(status, 0)
    score += {"low": 20, "medium": 5, "high": -15, "unknown": 0}.get(risk, 0)
    score += (_num(supplier.get("confiance_note"), 0) / 5) * 6
    score += (_num(supplier.get("service_note"), 0) / 5) * 3
    score += (_num(supplier.get("contenu_note"), 0) / 5) * 3
    score += (_num(supplier.get("potentiel_negociation_note"), 0) / 5) * 3
    if supplier.get("paypal_gs") == "oui":
        score += 8
    if supplier.get("tracking") == "oui":
        score += 4
    if supplier.get("photos_video") == "oui":
        score += 3
    elif supplier.get("photos_video") == "partiel":
        score += 1
    landed = supplier.get("landed_cost_per_card_eur")
    purchase_pct = supplier.get("purchase_percentage_of_resale")
    margin_pct = supplier.get("estimated_margin_percent")
    if landed is not None:
        score += 12
        if landed <= 1.8:
            score += 12
        elif landed <= 2.1:
            score += 7
        elif landed <= 2.4:
            score += 3
    elif calculate_offer(supplier, 200).get("available") or calculate_offer(supplier, 100).get("available"):
        score += 3
    if purchase_pct is not None:
        if purchase_pct <= 55:
            score += 10
        elif purchase_pct <= 65:
            score += 6
        elif purchase_pct <= 75:
            score += 2
    if margin_pct is not None:
        if margin_pct >= 45:
            score += 8
        elif margin_pct >= 35:
            score += 5
        elif margin_pct >= 25:
            score += 2
    decision = _display_action_short(supplier.get("action_recommandee"))
    decision_scores = {
        "Demander un devis complet": 4,
        "Continuer la négociation": 6,
        "Garder comme solution de secours": -6,
        "Ne pas prioriser": -25,
        "Mettre la discussion en pause": -20,
        "Éviter ce fournisseur": -40,
    }
    score += decision_scores.get(decision, 0)
    if status in {"à éviter", "archivé"}:
        score = -100
    if supplier.get("doublons") == "peu":
        score += 3
    elif supplier.get("doublons") == "beaucoup":
        score -= 6
    if risk == "high":
        score = min(score, 42)
    return max(-100, min(100, round(score, 1)))


def _priority_label(score):
    if score >= 70:
        return "Priorité haute"
    if score >= 50:
        return "À suivre"
    if score >= 30:
        return "Secondaire"
    if score >= 1:
        return "Faible priorité"
    return "À éviter"


def _priority_badge(score):
    label = _priority_label(score)
    tone = {
        "Priorité haute": "success",
        "À suivre": "info",
        "Secondaire": "neutral",
        "Faible priorité": "neutral",
        "À éviter": "danger",
    }.get(label, "neutral")
    return _badge(label, tone)


def _display_action(action):
    if not action:
        return "—"
    text = str(action).strip()
    translations = {
        "Ask for 100 and 200-card quotes, shipping, payment, condition, duplicates and sample photos.": "Demander les prix pour 100 et 200 cartes, le port, le paiement, l’état, les doublons et des photos d’exemple.",
        "Wait and reconnect later when they have more time or suitable AR stock.": "Attendre puis reprendre contact lorsqu’il aura plus de temps ou du stock AR adapté.",
        "Wait for response and ask whether future AR restocks or bulk lots are possible.": "Attendre sa réponse puis demander si de futurs réassorts AR ou lots en quantité sont possibles.",
        "Do not prioritise.": "Ne pas prioriser ce fournisseur.",
        "Do not continue sourcing discussion for now.": "Mettre la discussion en pause pour le moment.",
        "Only follow up if they can provide a concrete AR/CHR quote.": "Relancer uniquement s’il peut proposer un prix précis pour un lot AR/CHR.",
        "Request exact final PayPal quote and discount confirmation.": "Demander un devis PayPal final et la confirmation d’une remise.",
        "Request a detailed 100-card and 200-card Japanese AR/CHR quote, shipping to France, PayPal G&S availability, condition, duplicate policy and sample evidence.": "Demander un devis détaillé pour 100 et 200 cartes AR/CHR japonaises, avec port vers la France, PayPal G&S, état, doublons et photos d’exemple.",
        "Keep as backup. Confirm whether the $310 total includes PayPal G&S before any decision.": "Garder comme solution de secours et vérifier si les 310 $ incluent PayPal G&S.",
        "Keep as backup only.": "Garder uniquement comme solution de secours.",
        "Request a detailed quote for 200 cards, including shipping to France and PayPal Goods & Services.": "Demander un devis détaillé pour 200 cartes, avec port vers la France et PayPal G&S.",
        "Request a 200-card AR/CHR quote with shipping to France, PayPal G&S fee, condition, duplicate policy and sample evidence.": "Demander un devis AR/CHR pour 200 cartes, avec port vers la France, frais PayPal G&S, état, doublons et photos d’exemple.",
        "Send clear polite decline. No further negotiation unless supplier offers a substantially better condition.": "Refuser poliment. Ne reprendre contact que si les conditions deviennent nettement meilleures.",
        "Avoid unless pricing changes significantly.": "Éviter ce fournisseur sauf si le prix baisse fortement.",
        "Keep as an alternative only if custom preferences become important.": "Garder seulement comme alternative si des besoins spécifiques deviennent importants.",
    }
    normalized = " ".join(text.rstrip(".").lower().split())
    normalized_map = {" ".join(key.rstrip(".").lower().split()): value for key, value in translations.items()}
    if normalized in normalized_map:
        return normalized_map[normalized]
    if all(ord(char) < 128 for char in text) and any(word in text.lower() for word in ["ask", "wait", "follow", "continue", "prioritise", "prioritize"]):
        return "Action à consulter dans le détail"
    return text


def _display_action_short(action):
    full = _display_action(action)
    if full in {"—", "Action à consulter dans le détail"}:
        return full
    low = full.lower()
    if "éviter" in low or "refuser" in low:
        return "Éviter ce fournisseur"
    if "devis" in low or "prix" in low:
        return "Demander un devis complet"
    if "négociation" in low or "remise" in low:
        return "Continuer la négociation"
    if "attendre" in low:
        return "Attendre la réponse"
    if "secours" in low or "alternative" in low:
        return "Garder comme solution de secours"
    if "prioriser" in low:
        return "Ne pas prioriser"
    if "pause" in low:
        return "Mettre la discussion en pause"
    return full if len(full) <= 60 else "Action à consulter dans le détail"


def _supplier_status_line(supplier):
    return " · ".join([
        _clean_text(supplier.get("pays"), "Pays inconnu"),
        _clean_text(supplier.get("type"), "Autre"),
        _clean_text(supplier.get("statut"), "Non renseigné"),
    ])


def _render_supplier_summary_card(supplier, suppliers):
    score_info = pokestock_score(supplier, suppliers)
    conf = confidence_score(supplier)
    priority = _priority_score(supplier)
    badges = " · ".join(_supplier_badge_labels(supplier, limit=3)) or "Aucun badge"
    action = _display_action(supplier.get("action_recommandee"))
    with st.container(border=True):
        head, open_col = st.columns([4, 1])
        with head:
            st.markdown(f"#### {supplier.get('nom') or 'Sans nom'}")
            st.caption(_supplier_status_line(supplier))
        with open_col:
            st.caption("Voir le fournisseur")
        st.markdown(_priority_badge(priority), unsafe_allow_html=True)
        price_col, score_col, risk_col, trust_col = st.columns(4)
        price_col.metric("Prix / carte", _supplier_unit_price_label(supplier))
        price_col.caption(_supplier_price_hint(supplier))
        score_col.metric(
            "Score PokéStock",
            f"{score_info['score']:.0f}/100" if score_info.get("score") is not None else "Incomplet",
        )
        risk_col.metric("Risque", _risk_label(supplier.get("risk_level")))
        trust_col.metric("Confiance", f"{conf:.1f}/5" if conf is not None else "—")
        st.caption(badges)
        st.markdown("**Décision recommandée**")
        st.caption(action if len(action) <= 220 else action[:217].rstrip() + "...")
        _render_test_order_sheet(supplier, {"resale_scenario": "reference", "custom_resale_per_card_eur": 3.0})


def _supplier_badge_labels(supplier, limit=3):
    labels = []
    potential = supplier.get("potentiel_negociation_note")
    if potential is not None and potential >= 4:
        labels.append("Très flexible")
    elif potential is not None and 2 <= potential < 4 and supplier.get("statut") not in {"archivé", "à éviter"}:
        labels.append("À négocier")
    if supplier.get("paypal_gs") == "oui":
        labels.append("PayPal G&S")
    if supplier.get("tracking") == "oui":
        labels.append("Tracking")
    if supplier.get("photos_video") == "oui":
        labels.append("Photos")
    elif supplier.get("photos_video") == "partiel":
        labels.append("Photos partielles")
    if supplier.get("devise") != "EUR" and not supplier.get("conversion_rate_to_eur"):
        labels.append("Devise à convertir")
    status = supplier.get("statut")
    if status == "top fournisseur":
        labels.append("Top fournisseur")
    elif status == "à éviter":
        labels.append("À éviter")
    elif status == "testé":
        labels.append("Testé")
    return labels[:limit]


def _supplier_badges(supplier):
    badges = []
    potential = supplier.get("potentiel_negociation_note")
    if potential is not None and potential >= 4:
        badges.append(_badge("Très flexible", "success"))
    elif potential is not None and 2 <= potential < 4 and supplier.get("statut") not in {"archivé", "à éviter"}:
        badges.append(_badge("À négocier", "warning"))
    if supplier.get("paypal_gs") == "oui":
        badges.append(_badge("PayPal G&S", "info"))
    if supplier.get("photos_video") == "oui":
        badges.append(_badge("Photos / vidéo", "success"))
    elif supplier.get("photos_video") == "partiel":
        badges.append(_badge("Photos partielles", "warning"))
    if supplier.get("tracking") == "oui":
        badges.append(_badge("Tracking", "info"))
    if supplier.get("devise") != "EUR" and not supplier.get("conversion_rate_to_eur"):
        badges.append(_badge("Devise à convertir", "warning"))
    status = supplier.get("statut")
    if status == "top fournisseur":
        badges.append(_badge("Top fournisseur", "success"))
    elif status == "à éviter":
        badges.append(_badge("À éviter", "danger"))
    elif status == "testé":
        badges.append(_badge("Testé", "info"))
    elif status == "à contacter":
        badges.append(_badge("À contacter", "neutral"))
    elif status == "en attente":
        badges.append(_badge("En attente", "warning"))
    return " ".join(badges[:3])


def _import_format_label(import_format):
    return {
        "supplier_review_v1": "Review structurée détectée",
        "supplier_update": "Mise à jour fournisseur détectée",
        "free_text": "Texte libre détecté",
    }.get(import_format, "Texte libre détecté")


def _field_label(field):
    labels = {
        "nom": "Nom",
        "pays": "Pays",
        "type": "Type",
        "statut": "Statut",
        "prix_100": "Prix 100",
        "prix_200": "Prix 200",
        "devise": "Devise",
        "port": "Port",
        "paiement": "Paiement",
        "paypal_gs": "PayPal G&S",
        "frais_paypal": "Frais PayPal",
        "tracking": "Tracking",
        "photos_video": "Photos / vidéo",
        "etat": "État",
        "doublons": "Doublons",
        "confiance_note": "Confiance",
        "service_note": "Service",
        "contenu_note": "Contenu",
        "potentiel_negociation_note": "Potentiel de négociation",
        "derniere_analyse": "Dernière analyse",
        "action_recommandee": "Action recommandée",
        "market_position": "Position marché",
        "market_position_rank": "Rang position marché",
        "best_for": "Idéal pour",
        "best_for_tags": "Tags idéal pour",
        "risk_level": "Niveau de risque",
        "recommended_first_order": "Première commande recommandée",
        "recommended_first_order_quantity": "Quantité première commande",
        "review_last_updated": "Last Updated review",
        "pros": "Pros",
        "cons": "Cons",
        "notes": "Notes",
    }
    return labels.get(field, field)


def _display_value(value):
    if isinstance(value, list):
        return ", ".join(str(item) for item in value) or "—"
    if value in (None, "", []):
        return "—"
    return str(value)


def _risk_badge(risk_level):
    mapping = {
        "low": ("Risque faible", "success"),
        "medium": ("Risque moyen", "warning"),
        "high": ("Risque élevé", "danger"),
        "unknown": ("Risque non renseigné", "neutral"),
    }
    label, tone = mapping.get(risk_level or "unknown", mapping["unknown"])
    return _badge(label, tone)


def _format_review_date(value):
    if not value:
        return ""
    try:
        year, month, day = str(value).split("-", 2)
        return f"{day}/{month}/{year}"
    except ValueError:
        return ""


def _candidate_differences(existing, candidate, detected_fields):
    rows = []
    for field in detected_fields:
        if field in {"notes", "analysis_updated_at"}:
            continue
        old = (existing or {}).get(field)
        new = candidate.get(field)
        if _display_value(old) != _display_value(new):
            rows.append({"Champ": _field_label(field), "Avant": _display_value(old), "Nouveau": _display_value(new)})
    return rows


def _merge_detected_fields(existing, candidate, detected_fields):
    merged = dict(existing)
    for field in detected_fields:
        if field in {"id", "created_at", "updated_at"}:
            continue
        merged[field] = candidate.get(field)
    merged["id"] = existing.get("id")
    merged["created_at"] = existing.get("created_at")
    return normalize_supplier(merged)


def _with_price_history(payload, entry):
    if entry:
        payload = dict(payload)
        payload.setdefault("price_history", [])
        payload["price_history"] = list(payload["price_history"]) + [entry]
    return payload


def _supplier_score_text(supplier, suppliers):
    score = pokestock_score(supplier, suppliers)
    return f"{score['score']:.0f}/100" if score.get("score") is not None else "Incomplet"


def _comparison_rows(suppliers):
    rows = []
    for supplier in suppliers:
        score = pokestock_score(supplier, suppliers)
        rows.append({
            "Fournisseur": supplier.get("nom") or "Sans nom",
            "Prix / carte EUR": supplier.get("prix_unitaire_eur"),
            "Score PokéStock": score.get("score"),
            "Couverture": score.get("coverage"),
            "Risque": supplier.get("risk_level", "unknown"),
            "Confiance": supplier.get("confiance_note"),
            "Potentiel négo": supplier.get("potentiel_negociation_note"),
            "Badges": " ".join((supplier.get("best_for_tags") or [])[:2]),
            "Action recommandée": supplier.get("action_recommandee") or "",
            "Pays": supplier.get("pays") or "",
            "Type": supplier.get("type") or "",
            "Statut": supplier.get("statut") or "",
            "Prix 100": supplier.get("prix_100"),
            "Prix 200": supplier.get("prix_200"),
            "Devise": supplier.get("devise") or "",
            "Port": supplier.get("port"),
            "PayPal G&S": supplier.get("paypal_gs"),
            "Tracking": supplier.get("tracking"),
            "Photos / vidéo": supplier.get("photos_video"),
            "Best For": supplier.get("best_for") or "",
            "Position marché": supplier.get("market_position") or "",
            "Première commande": supplier.get("recommended_first_order") or "",
            "Dernière mise à jour": supplier.get("review_last_updated") or "",
        })
    return rows


def _comparison_rows(suppliers):
    rows = []
    for supplier in suppliers:
        score = pokestock_score(supplier, suppliers)
        priority = _priority_score(supplier)
        rows.append({
            "Fournisseur": supplier.get("nom") or "Sans nom",
            "Coût rendu / carte": _landed_cost_label(supplier),
            "% achat / cote": _format_percent_range(
                supplier.get("purchase_percentage_of_resale_min"),
                supplier.get("purchase_percentage_of_resale_max"),
                supplier.get("purchase_percentage_of_resale"),
            ),
            "Marge estimée": _margin_label(supplier),
            "Prix cartes €": _eur_source_label(supplier.get("cards_price_eur"), supplier.get("cards_price_source"), supplier.get("cards_price_source_currency")),
            "Livraison €": _eur_source_label(supplier.get("shipping_price_eur"), supplier.get("shipping_price_source"), supplier.get("shipping_price_source_currency"), "À confirmer"),
            "PayPal €": _eur_source_label(supplier.get("paypal_fee_eur"), supplier.get("paypal_fee_source"), supplier.get("paypal_fee_source_currency"), "À confirmer"),
            "Total fournisseur €": _eur_source_label(supplier.get("supplier_total_eur"), supplier.get("supplier_total_source"), supplier.get("supplier_total_source_currency"), "Incomplet"),
            "Prix cartes": _cards_price_label(supplier),
            "Livraison": _shipping_label(supplier),
            "PayPal": _paypal_label(supplier),
            "Total fournisseur": _supplier_total_label(supplier),
            "Prix / carte": _supplier_unit_price_label(supplier),
            "Priorité": _priority_label(priority),
            "Indice priorité": priority,
            "Score": f"{score.get('score'):.0f}/100" if score.get("score") is not None else "Incomplet",
            "Couverture": score.get("coverage"),
            "Risque": _risk_label(supplier.get("risk_level")),
            "Confiance": f"{supplier.get('confiance_note'):.1f}/5" if supplier.get("confiance_note") is not None else "—",
            "Potentiel négo": f"{supplier.get('potentiel_negociation_note'):.1f}/5" if supplier.get("potentiel_negociation_note") is not None else "—",
            "Badges": " · ".join(_supplier_badge_labels(supplier, limit=2)) or "—",
            "Idéal pour": ", ".join(supplier.get("best_for_tags") or []) or "—",
            "Décision": _display_action_short(supplier.get("action_recommandee")),
            "Action recommandée": _display_action(supplier.get("action_recommandee")),
            "Pays": _clean_text(supplier.get("pays"), ""),
            "Type": _clean_text(supplier.get("type"), ""),
            "Statut": _clean_text(supplier.get("statut"), ""),
            "Prix 100": supplier.get("prix_100"),
            "Prix 200": supplier.get("prix_200"),
            "Devise": supplier.get("devise") or "",
            "Rendu France min": supplier.get("landed_cost_estimated_min_eur"),
            "Rendu France max": supplier.get("landed_cost_estimated_max_eur"),
            "Revente estimée": supplier.get("estimated_resale_total_eur"),
            "Port": supplier.get("port"),
            "PayPal G&S": _clean_text(supplier.get("paypal_gs")),
            "Tracking": _clean_text(supplier.get("tracking")),
            "Photos / vidéo": _clean_text(supplier.get("photos_video")),
            "Best For": supplier.get("best_for") or "",
            "Position marché": supplier.get("market_position") or "",
            "Première commande": supplier.get("recommended_first_order") or "",
            "Dernière mise à jour": supplier.get("review_last_updated") or "",
        })
    return rows


def _build_suppliers_workbook(suppliers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparatif"
    headers = [
        "Fournisseur", "Pays", "Type", "Statut", "Prix cartes", "Livraison", "Frais PayPal",
        "Total fournisseur", "Devise", "Rendu France min", "Rendu France max",
        "Coût rendu / carte min", "Coût rendu / carte max", "Revente estimée",
        "% achat / cote min", "% achat / cote max", "Marge estimée min", "Marge estimée max",
        "Confiance", "Service", "Contenu", "Score PokéStock", "Couverture données", "Risque",
        "Potentiel négociation", "Best For", "Position marché", "Première commande recommandée",
        "PayPal G&S", "Tracking", "Photos / vidéo", "Action recommandée", "Dernière analyse",
        "Dernière mise à jour review",
    ]
    ws.append(headers)
    for supplier in suppliers:
        score = pokestock_score(supplier, suppliers)
        ws.append([
            supplier.get("nom"), supplier.get("pays"), supplier.get("type"), supplier.get("statut"),
            supplier.get("cards_price_source"), supplier.get("shipping_price_source"), supplier.get("paypal_fee_source"),
            supplier.get("supplier_total_source"), supplier.get("supplier_total_source_currency") or supplier.get("devise"),
            supplier.get("landed_cost_estimated_min_eur"), supplier.get("landed_cost_estimated_max_eur"),
            supplier.get("landed_cost_per_card_min_eur"), supplier.get("landed_cost_per_card_max_eur"),
            supplier.get("estimated_resale_total_eur"), supplier.get("purchase_percentage_of_resale_min"),
            supplier.get("purchase_percentage_of_resale_max"), supplier.get("estimated_margin_min_eur"),
            supplier.get("estimated_margin_max_eur"), supplier.get("confiance_note"), supplier.get("service_note"),
            supplier.get("contenu_note"), score.get("score"), score.get("coverage"), supplier.get("risk_level"),
            supplier.get("potentiel_negociation_note"), supplier.get("best_for"), supplier.get("market_position"),
            supplier.get("recommended_first_order"), supplier.get("paypal_gs"), supplier.get("tracking"),
            supplier.get("photos_video"), supplier.get("action_recommandee"), supplier.get("derniere_analyse"),
            supplier.get("review_last_updated"),
        ])
    hist = wb.create_sheet("Historique offres")
    hist.append(["Fournisseur", "Date", "Ancien prix", "Nouveau prix", "Devise", "Quantité", "Ancien prix unitaire", "Nouveau prix unitaire", "Analyse", "Action recommandée"])
    for supplier in suppliers:
        for entry in supplier.get("price_history", []):
            hist.append([
                supplier.get("nom"), entry.get("date"), entry.get("ancien_prix"), entry.get("nouveau_prix"),
                entry.get("devise"), entry.get("quantite"), entry.get("ancien_prix_unitaire"),
                entry.get("nouveau_prix_unitaire"), entry.get("analyse"), entry.get("action_recommandee"),
            ])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def _supplier_from_excel_row(row):
    supplier = default_supplier()
    mapping = {
        "Fournisseur": "nom", "Pays": "pays", "Type": "type", "Statut": "statut",
        "Prix 100": "prix_100", "Prix 200": "prix_200", "Devise": "devise", "Port": "port",
        "Prix cartes": "cards_price_source", "Livraison": "shipping_price_source",
        "Frais PayPal": "paypal_fee_source", "Total fournisseur": "supplier_total_source",
        "Rendu France min": "landed_cost_estimated_min_eur", "Rendu France max": "landed_cost_estimated_max_eur",
        "Coût rendu / carte min": "landed_cost_per_card_min_eur", "Coût rendu / carte max": "landed_cost_per_card_max_eur",
        "Revente estimée": "estimated_resale_total_eur",
        "% achat / cote min": "purchase_percentage_of_resale_min", "% achat / cote max": "purchase_percentage_of_resale_max",
        "Marge estimée min": "estimated_margin_min_eur", "Marge estimée max": "estimated_margin_max_eur",
        "PayPal G&S": "paypal_gs", "Tracking": "tracking", "Photos / vidéo": "photos_video",
        "Confiance": "confiance_note", "Service": "service_note", "Contenu": "contenu_note",
        "Risque": "risk_level", "Action recommandée": "action_recommandee",
        "Best For": "best_for", "Position marché": "market_position",
        "Première commande recommandée": "recommended_first_order",
        "Dernière mise à jour review": "review_last_updated",
    }
    detected = []
    for header, field in mapping.items():
        value = row.get(header)
        if value in (None, ""):
            continue
        supplier[field] = value
        detected.append(field)
    supplier["market_position_rank"] = _parse_market_position_rank(supplier.get("market_position"))
    supplier["best_for_tags"] = _parse_best_for_tags(supplier.get("best_for"))
    supplier["recommended_first_order_quantity"] = _parse_first_order_quantity(supplier.get("recommended_first_order"))
    supplier["review_last_updated"] = _parse_iso_date(supplier.get("review_last_updated")) or ""
    currency = supplier.get("devise") or "EUR"
    for field in ["cards_price_source_currency", "shipping_price_source_currency", "paypal_fee_source_currency", "supplier_total_source_currency"]:
        supplier[field] = currency
    return normalize_supplier(supplier), detected


def _parse_excel_suppliers(uploaded_file):
    wb = load_workbook(uploaded_file, data_only=True)
    sheet = wb["Comparatif"] if "Comparatif" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    parsed = []
    for values in rows[1:]:
        row = {headers[idx]: value for idx, value in enumerate(values) if idx < len(headers) and headers[idx]}
        supplier, detected = _supplier_from_excel_row(row)
        if supplier.get("nom"):
            parsed.append({"supplier": supplier, "detected_fields": detected})
    return parsed


def _option_index(options, value):
    try:
        return options.index(value)
    except ValueError:
        return 0


def _supplier_form(prefix, supplier=None, submit_label="Enregistrer"):
    supplier = normalize_supplier(supplier or default_supplier())
    with st.form(f"{prefix}_form"):
        st.markdown("**Identité et contact**")
        c1, c2 = st.columns(2)
        nom = c1.text_input("Nom fournisseur", value=supplier.get("nom", ""), key=f"{prefix}_nom")
        pays = c2.text_input("Pays", value=supplier.get("pays", ""), key=f"{prefix}_pays")
        contact_link = st.text_input("Contact / lien", value=supplier.get("contact_link", ""), key=f"{prefix}_contact")
        type_value = st.selectbox("Type", SUPPLIER_TYPES, index=_option_index(SUPPLIER_TYPES, supplier.get("type")), key=f"{prefix}_type")

        st.markdown("**Prix et devis**")
        p1, p2, p3 = st.columns(3)
        prix_100 = p1.number_input("Prix pour 100 cartes", min_value=0.0, value=_num(supplier.get("prix_100")), step=1.0, key=f"{prefix}_prix100")
        prix_200 = p2.number_input("Prix pour 200 cartes", min_value=0.0, value=_num(supplier.get("prix_200")), step=1.0, key=f"{prefix}_prix200")
        devise = p3.selectbox("Devise", SUPPLIER_CURRENCIES, index=_option_index(SUPPLIER_CURRENCIES, supplier.get("devise")), key=f"{prefix}_devise")
        p4, p5, p6 = st.columns(3)
        port = p4.number_input("Port", min_value=0.0, value=_num(supplier.get("port")), step=1.0, key=f"{prefix}_port")
        paiement = p5.text_input("Paiement", value=supplier.get("paiement", ""), key=f"{prefix}_paiement")
        paypal_gs = p6.selectbox("PayPal G&S", PAYPAL_STATES, index=_option_index(PAYPAL_STATES, supplier.get("paypal_gs")), key=f"{prefix}_paypal")
        frais_paypal = st.number_input("Frais PayPal", min_value=0.0, value=_num(supplier.get("frais_paypal")), step=1.0, key=f"{prefix}_paypal_fees")

        st.markdown("**Import et coûts additionnels**")
        i1, i2 = st.columns(2)
        tva_import_active = i1.checkbox("TVA / import activé", value=bool(supplier.get("tva_import_active")), key=f"{prefix}_tva_active")
        tva_import_rate_pct = i2.number_input("Taux TVA / import estimé (%)", min_value=0.0, max_value=100.0, value=_num(supplier.get("tva_import_rate")) * 100, step=0.5, key=f"{prefix}_tva_rate")
        i3, i4 = st.columns(2)
        frais_import = i3.number_input("Frais import", min_value=0.0, value=_num(supplier.get("frais_import")), step=1.0, key=f"{prefix}_import")
        frais_dossier = i4.number_input("Frais de dossier", min_value=0.0, value=_num(supplier.get("frais_dossier")), step=1.0, key=f"{prefix}_dossier")

        st.markdown("**Informations sur les cartes**")
        m1, m2, m3, m4 = st.columns(4)
        etat = m1.selectbox("État", CARD_STATES, index=_option_index(CARD_STATES, supplier.get("etat")), key=f"{prefix}_etat")
        doublons = m2.selectbox("Doublons", DUPLICATE_STATES, index=_option_index(DUPLICATE_STATES, supplier.get("doublons")), key=f"{prefix}_doublons")
        photos_video = m3.selectbox("Photos / vidéo", MEDIA_STATES, index=_option_index(MEDIA_STATES, supplier.get("photos_video")), key=f"{prefix}_photos")
        tracking = m4.selectbox("Tracking", TRACKING_STATES, index=_option_index(TRACKING_STATES, supplier.get("tracking")), key=f"{prefix}_tracking")
        valeur_lot_estimee = st.number_input("Valeur de lot estimée", min_value=0.0, value=_num(supplier.get("valeur_lot_estimee")), step=1.0, key=f"{prefix}_value")

        st.markdown("**Notes et statut**")
        n1, n2, n3, n4 = st.columns(4)
        confiance_note = n1.slider("Confiance", 0.0, 5.0, _num(supplier.get("confiance_note")), 0.5, key=f"{prefix}_trust")
        service_note = n2.slider("Service", 0.0, 5.0, _num(supplier.get("service_note")), 0.5, key=f"{prefix}_service")
        contenu_note = n3.slider("Contenu", 0.0, 5.0, _num(supplier.get("contenu_note")), 0.5, key=f"{prefix}_content")
        statut = n4.selectbox("Statut", SUPPLIER_STATUSES, index=_option_index(SUPPLIER_STATUSES, supplier.get("statut")), key=f"{prefix}_status")
        potential_cols = st.columns(2)
        potentiel_negociation_note = potential_cols[0].slider("Potentiel de négociation", 0.0, 5.0, _num(supplier.get("potentiel_negociation_note")), 0.5, key=f"{prefix}_negotiation_potential")
        negotiation_note = potential_cols[1].text_input("Note négociation", value=supplier.get("negotiation_note", ""), key=f"{prefix}_negotiation_note")
        specialites = st.multiselect("Spécialités", SUPPLIER_SPECIALTIES, default=supplier.get("specialites") or [], key=f"{prefix}_specialites")
        q1, q2 = st.columns(2)
        grosses_quantites = q1.selectbox(
            "Grosses quantités confirmées",
            BIG_QUANTITY_STATES,
            index=_option_index(BIG_QUANTITY_STATES, supplier.get("grosses_quantites_confirmees")),
            key=f"{prefix}_big_qty",
        )
        quantite_minimum = q2.number_input(
            "Quantité minimum",
            min_value=0,
            value=int(supplier.get("quantite_minimum") or 0),
            step=1,
            key=f"{prefix}_min_qty",
        )
        conversion_rate = 1.0
        if devise != "EUR":
            conversion_rate = st.number_input("Taux de conversion vers EUR", min_value=0.0, value=_num(supplier.get("conversion_rate_to_eur")), step=0.0001, format="%.6f", key=f"{prefix}_conversion")
        analysis_cols = st.columns(2)
        derniere_analyse = analysis_cols[0].text_area("Dernière analyse", value=supplier.get("derniere_analyse", ""), key=f"{prefix}_analysis")
        action_recommandee = analysis_cols[1].text_area("Action recommandée", value=supplier.get("action_recommandee", ""), key=f"{prefix}_recommended_action")
        st.markdown("**Métadonnées de review**")
        meta_cols = st.columns(3)
        market_position = meta_cols[0].text_input("Position marché", value=supplier.get("market_position", ""), key=f"{prefix}_market_position")
        risk_options = ["unknown", "low", "medium", "high"]
        risk_level = meta_cols[1].selectbox(
            "Niveau de risque",
            risk_options,
            index=_option_index(risk_options, supplier.get("risk_level")),
            key=f"{prefix}_risk_level",
        )
        review_last_updated = meta_cols[2].text_input("Last Updated review", value=supplier.get("review_last_updated", ""), key=f"{prefix}_review_last_updated")
        best_for = st.text_input("Idéal pour", value=supplier.get("best_for", ""), key=f"{prefix}_best_for")
        recommended_first_order = st.text_input("Première commande recommandée", value=supplier.get("recommended_first_order", ""), key=f"{prefix}_recommended_first_order")
        pros_text = st.text_area("Points forts détectés / pros", value="\n".join(supplier.get("pros") or []), key=f"{prefix}_pros")
        cons_text = st.text_area("Points faibles détectés / cons", value="\n".join(supplier.get("cons") or []), key=f"{prefix}_cons")
        notes = st.text_area("Notes libres", value=supplier.get("notes", ""), key=f"{prefix}_notes")
        save_history = st.checkbox("Ajouter cette offre à l'historique", value=False, key=f"{prefix}_history")

        submitted = st.form_submit_button(submit_label, type="primary")

    if not submitted:
        return None
    updated = dict(supplier)
    analysis_updated_at = supplier.get("analysis_updated_at") or ""
    if (derniere_analyse.strip() != supplier.get("derniere_analyse", "")) or (action_recommandee.strip() != supplier.get("action_recommandee", "")):
        analysis_updated_at = now_iso()
    parsed_review_date = _parse_iso_date(review_last_updated)
    updated.update(
        {
            "nom": nom.strip(),
            "pays": pays.strip(),
            "contact_link": contact_link.strip(),
            "type": type_value,
            "prix_100": prix_100 if prix_100 > 0 else None,
            "prix_200": prix_200 if prix_200 > 0 else None,
            "devise": devise,
            "port": port,
            "paiement": paiement.strip(),
            "paypal_gs": paypal_gs,
            "frais_paypal": frais_paypal,
            "tva_import_active": tva_import_active,
            "tva_import_rate": tva_import_rate_pct / 100 if tva_import_active else 0.0,
            "frais_import": frais_import,
            "frais_dossier": frais_dossier,
            "etat": etat,
            "doublons": doublons,
            "photos_video": photos_video,
            "tracking": tracking,
            "valeur_lot_estimee": valeur_lot_estimee if valeur_lot_estimee > 0 else None,
            "conversion_rate_to_eur": 1.0 if devise == "EUR" else (conversion_rate if conversion_rate > 0 else None),
            "confiance_note": confiance_note,
            "service_note": service_note,
            "contenu_note": contenu_note,
            "potentiel_negociation_note": potentiel_negociation_note,
            "negotiation_note": negotiation_note.strip(),
            "pros": [line.strip() for line in pros_text.splitlines() if line.strip()],
            "cons": [line.strip() for line in cons_text.splitlines() if line.strip()],
            "derniere_analyse": derniere_analyse.strip(),
            "action_recommandee": action_recommandee.strip(),
            "analysis_updated_at": analysis_updated_at,
            "market_position": market_position.strip(),
            "market_position_rank": _parse_market_position_rank(market_position),
            "best_for": best_for.strip(),
            "best_for_tags": _parse_best_for_tags(best_for),
            "risk_level": risk_level,
            "recommended_first_order": recommended_first_order.strip(),
            "recommended_first_order_quantity": _parse_first_order_quantity(recommended_first_order),
            "follow_status": follow_status,
            "follow_up_date": follow_up_date.strip(),
            "follow_up_note": follow_up_note.strip(),
            "preferred_offer_variant": preferred_offer_variant.strip(),
            "review_last_updated": parsed_review_date or supplier.get("review_last_updated") or "",
            "notes": notes.strip(),
            "statut": statut,
            "specialites": specialites,
            "grosses_quantites_confirmees": grosses_quantites,
            "quantite_minimum": quantite_minimum if quantite_minimum > 0 else None,
            "_save_history": save_history,
        }
    )
    return normalize_supplier(updated)


def _render_calculation(supplier, quantity=200):
    offer = calculate_offer(supplier, quantity)
    if not offer.get("available"):
        st.caption(offer.get("reason", "Calcul indisponible"))
        return
    st.markdown(
        f"""
        <div class="supplier-calc">
            <div>Prix {quantity} : <strong>{_money(offer['base_price'], offer['currency'])}</strong></div>
            <div>Port : {_money(offer['port'], offer['currency'])}</div>
            <div>Frais PayPal : {_money(offer['frais_paypal'], offer['currency'])}</div>
            <div>TVA / import : {_money(offer['import_amount'], offer['currency'])}</div>
            <div>Frais import + dossier : {_money(offer['frais_import'] + offer['frais_dossier'], offer['currency'])}</div>
            <div class="supplier-calc-total">Final : {_money(offer['final'], offer['currency'])} · Unitaire : {_money(offer['unit'], offer['currency'])}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if offer.get("conversion_needed"):
        st.warning("Conversion nécessaire : ce fournisseur n'est pas inclus dans les classements prix.")
    elif offer.get("currency") != "EUR":
        st.caption(f"Taux manuel : {offer.get('conversion_rate'):.6f} EUR · Final EUR : {_money(offer.get('final_eur'))} · Unitaire EUR : {_money(offer.get('unit_eur'))}")


def _render_rankings(suppliers):
    rankings = supplier_rankings(suppliers)
    st.markdown("### Classements")
    st.caption("Classement indicatif basé sur le prix et vos notes personnelles.")
    c1, c2, c3 = st.columns(3)
    groups = [
        (c1, "Meilleur prix", rankings["best_price"], "landed_cost_per_card_eur"),
        (c2, "Meilleure confiance", rankings["best_confidence"], None),
        (c3, "Meilleur équilibre", rankings["best_balance"], None),
    ]
    for col, title, items, price_field in groups:
        with col:
            st.markdown(f"**{title}**")
            if not items:
                st.caption("Aucune donnée suffisante.")
                continue
            for item in items[:3]:
                score = confidence_score(item)
                detail = _money(item.get(price_field)) if price_field else (f"{score:.1f}/5" if score is not None else "—")
                st.markdown(f"- **{item.get('nom') or 'Sans nom'}** · {detail}")


def _render_score(supplier, suppliers):
    score = pokestock_score(supplier, suppliers)
    if score["complete"]:
        st.metric("Score PokéStock", f"{score['score']:.0f} / 100")
    else:
        st.metric("Score PokéStock", "Score incomplet")
    st.caption(f"Données couvertes : {score['coverage'] * 100:.0f} %")
    for key in ["prix", "confiance", "qualite", "paiement", "service"]:
        part = score["parts"].get(key, {})
        points = part.get("points")
        available = part.get("available", part.get("max"))
        if points is None:
            st.caption(f"{part.get('label', key)} : {part.get('reason', 'Non renseigné')}")
        else:
            st.caption(f"{part.get('label', key)} : {points:.1f} / {available:g} disponible(s)")
    if score["missing"]:
        st.caption("À compléter : " + ", ".join(score["missing"][:6]))


def _render_smart_rankings(suppliers):
    smart = smart_rankings(suppliers)
    st.markdown("### Classements PokéStock")
    st.caption("Classements indicatifs, basés sur vos données, non garantis.")
    labels = [
        ("Meilleur rapport qualité / prix", "best_value"),
        ("Fournisseur le plus fiable", "most_reliable"),
        ("Meilleur potentiel de négociation", "negotiable"),
        ("Meilleur pour AR mid/high", "ar_mid_high"),
        ("Meilleur pour débuter", "beginner"),
        ("Meilleur pour grosses quantités", "big_quantity"),
    ]
    cols = st.columns(3)
    for idx, (label, key) in enumerate(labels):
        with cols[idx % 3]:
            st.markdown(f"**{label}**")
            items = smart.get(key, [])
            if not items:
                st.caption("Historique ou données insuffisants.")
            else:
                for supplier in items[:3]:
                    st.caption(supplier.get("nom") or "Sans nom")


def _ranking_badges_for_supplier(supplier, suppliers):
    smart = smart_rankings(suppliers)
    badges = []
    eligible_content = [
        item for item in suppliers
        if item.get("contenu_note") is not None and item.get("statut") not in {"archivé", "à éviter"}
    ]
    eligible_content = sorted(eligible_content, key=lambda item: item.get("contenu_note") or -1, reverse=True)
    if smart.get("best_price") and smart["best_price"][0].get("id") == supplier.get("id"):
        badges.append(_badge("Meilleur prix", "success"))
    if eligible_content and eligible_content[0].get("id") == supplier.get("id"):
        badges.append(_badge("Meilleur contenu", "success"))
    mapping = [
        ("best_value", "Meilleur équilibre"),
        ("most_reliable", "Le plus fiable"),
        ("negotiable", "Négociable"),
        ("ar_mid_high", "AR mid/high"),
        ("beginner", "Débutant"),
        ("big_quantity", "Grosses quantités"),
    ]
    for key, label in mapping:
        items = smart.get(key, [])
        if items and items[0].get("id") == supplier.get("id"):
            badges.append(_badge(label, "info"))
            if len(badges) >= 3:
                break
    return " ".join(badges[:3])


def _event_label(event):
    actor = {"supplier": "Fournisseur", "me": "Toi", "system": "Système", "unknown": "Inconnu"}.get(event.get("actor"), "Inconnu")
    kind = {
        "supplier_offer": "Offre fournisseur",
        "my_offer": "Ton contre-projet",
        "agreement": "Accord",
        "refusal": "Refus",
        "negotiation_note": "Note",
    }.get(event.get("event_type"), "Note")
    amount = _money(event.get("amount"), event.get("currency")) if event.get("amount") is not None else ""
    return f"{event.get('event_date')} · {actor} · {kind} {amount}"


def _render_negotiations(data, supplier):
    stats = negotiation_stats(supplier)
    if stats.get("comparable_count"):
        st.success(
            f"{stats['comparable_count']} négociation(s) comparable(s) · "
            f"baisse moyenne : {stats['average_reduction_pct']:.1f} %"
            if stats.get("average_reduction_pct") is not None else "Négociations comparables enregistrées"
        )
    else:
        st.caption("Historique insuffisant pour mesurer une baisse accordée.")
    chart = negotiation_chart_series(supplier)
    if chart.get("available"):
        try:
            import plotly.graph_objects as go
            fig = go.Figure()
            for actor, name in [("supplier", "Offres fournisseur"), ("me", "Tes contre-offres")]:
                rows = [e for e in chart["events"] if e.get("actor") == actor]
                fig.add_trace(go.Scatter(
                    x=[e.get("event_date") for e in rows],
                    y=[e.get(chart["mode"]) for e in rows],
                    mode="lines+markers",
                    name=name,
                    text=[e.get("notes") for e in rows],
                ))
            fig.update_layout(height=280, margin=dict(l=8, r=8, t=18, b=8), paper_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(fig, width="stretch")
        except Exception:
            st.caption("Graphique indisponible dans cet environnement.")
    else:
        st.info(chart.get("reason"))

    with st.form(f"neg_add_{supplier['id']}"):
        st.markdown("**Ajouter une entrée de négociation**")
        c1, c2, c3 = st.columns(3)
        negotiation_id = c1.text_input("Négociation", value=f"neg-{supplier['id'][:6]}")
        actor = c2.selectbox("Acteur", NEGOTIATION_ACTORS)
        event_type = c3.selectbox("Type", NEGOTIATION_EVENT_TYPES)
        c4, c5, c6 = st.columns(3)
        quantity = c4.number_input("Quantité", min_value=0, value=200, step=1)
        amount = c5.number_input("Montant", min_value=0.0, value=0.0, step=1.0)
        currency = c6.selectbox("Devise", SUPPLIER_CURRENCIES)
        conversion = st.number_input("Taux vers EUR si nécessaire", min_value=0.0, value=0.0, step=0.0001, format="%.6f")
        payment_terms = st.text_input("Conditions de paiement")
        notes = st.text_area("Notes")
        if st.form_submit_button("Ajouter l'entrée", type="primary"):
            event = normalize_negotiation_event({
                "negotiation_id": negotiation_id,
                "actor": actor,
                "event_type": event_type,
                "quantity": quantity or None,
                "amount": amount if amount > 0 else None,
                "currency": currency,
                "conversion_rate_to_eur": conversion if conversion > 0 else None,
                "payment_terms": payment_terms,
                "notes": notes,
            })
            save_negotiation_event(data, supplier["id"], event)
            st.success("Entrée de négociation ajoutée.")
            st.rerun()
    for event in sorted(supplier.get("negotiation_history", []), key=lambda e: (e.get("event_date") or "", e.get("created_at") or ""), reverse=True):
        st.caption(_event_label(event))
        key = f"del_neg_{supplier['id']}_{event['id']}"
        if st.session_state.get(key):
            c1, c2 = st.columns(2)
            if c1.button("Confirmer suppression", key=f"confirm_{key}"):
                delete_negotiation_event(data, supplier["id"], event["id"])
                st.session_state.pop(key, None)
                st.rerun()
            if c2.button("Annuler", key=f"cancel_{key}"):
                st.session_state.pop(key, None)
                st.rerun()
        elif st.button("Supprimer cette entrée", key=key):
            st.session_state[key] = True
            st.rerun()


def _render_conversation(data, supplier):
    tag_filter = st.selectbox("Filtrer par tag", ["Tous"] + CONVERSATION_TAGS, key=f"tag_filter_{supplier['id']}")
    query = st.text_input("Rechercher dans les messages", key=f"msg_query_{supplier['id']}")
    messages = supplier.get("conversation_history", [])
    if tag_filter != "Tous":
        messages = [m for m in messages if tag_filter in (m.get("tags") or [])]
    if query:
        messages = [m for m in messages if query.lower() in (m.get("content") or "").lower()]
    with st.form(f"msg_add_{supplier['id']}"):
        st.markdown("**Ajouter un message local**")
        author = st.selectbox("Auteur", MESSAGE_AUTHORS)
        content = st.text_area("Message")
        tags = st.multiselect("Tags", CONVERSATION_TAGS)
        linked_neg = st.text_input("Lier à une négociation")
        linked_offer = st.text_input("Lier à une offre")
        if st.form_submit_button("Ajouter le message", type="primary"):
            save_conversation_message(data, supplier["id"], normalize_conversation_message({
                "author": author,
                "content": content,
                "tags": tags,
                "linked_negotiation_id": linked_neg,
                "linked_offer_history_id": linked_offer,
            }))
            st.success("Message ajouté.")
            st.rerun()
    with st.expander("Importer une conversation", expanded=False):
        st.caption("Local uniquement. Ne collez pas de mots de passe, coordonnées bancaires ou informations sensibles.")
        raw = st.text_area("Conversation à analyser", key=f"conv_import_{supplier['id']}")
        if st.button("Analyser la conversation", key=f"conv_analyze_{supplier['id']}"):
            st.session_state[f"conv_preview_{supplier['id']}"] = extract_conversation_from_text(raw)
        preview = st.session_state.get(f"conv_preview_{supplier['id']}")
        if preview:
            st.caption(f"{len(preview['messages'])} message(s) détecté(s), {len(preview['negotiation_events'])} offre(s) chiffrée(s) détectée(s).")
            if st.button("Importer les messages détectés", key=f"conv_save_{supplier['id']}"):
                for msg in preview["messages"]:
                    save_conversation_message(data, supplier["id"], msg)
                for event in preview["negotiation_events"]:
                    save_negotiation_event(data, supplier["id"], event)
                st.session_state.pop(f"conv_preview_{supplier['id']}", None)
                st.success("Conversation importée après confirmation.")
                st.rerun()
    for msg in sorted(messages, key=lambda m: (m.get("message_date") or "", m.get("created_at") or ""), reverse=True):
        st.markdown(f"**{msg.get('message_date')} · {msg.get('author')}**")
        st.caption(_clean_text(msg.get("content"), ""))
        if msg.get("tags"):
            st.caption("Tags : " + ", ".join(msg["tags"]))
        key = f"del_msg_{supplier['id']}_{msg['id']}"
        if st.session_state.get(key):
            c1, c2 = st.columns(2)
            if c1.button("Confirmer suppression", key=f"confirm_{key}"):
                delete_conversation_message(data, supplier["id"], msg["id"])
                st.session_state.pop(key, None)
                st.rerun()
            if c2.button("Annuler", key=f"cancel_{key}"):
                st.session_state.pop(key, None)
                st.rerun()
        elif st.button("Supprimer ce message", key=key):
            st.session_state[key] = True
            st.rerun()


def _render_compare(suppliers):
    if len(suppliers) < 2:
        st.caption("Ajoute au moins deux fournisseurs pour comparer.")
        return
    labels = [f"{s.get('nom') or 'Sans nom'} · {s.get('id')}" for s in suppliers]
    c1, c2 = st.columns(2)
    a_label = c1.selectbox("Fournisseur A", labels, key="compare_a")
    b_label = c2.selectbox("Fournisseur B", labels, index=1 if len(labels) > 1 else 0, key="compare_b")
    a = next(s for s in suppliers if s.get("id") == a_label.rsplit(" · ", 1)[-1])
    b = next(s for s in suppliers if s.get("id") == b_label.rsplit(" · ", 1)[-1])
    scores = {item["id"]: pokestock_score(item, suppliers) for item in (a, b)}
    rows = [
        ("Prix 100", _money(a.get("prix_100"), a.get("devise")), _money(b.get("prix_100"), b.get("devise"))),
        ("Prix 200", _money(a.get("prix_200"), a.get("devise")), _money(b.get("prix_200"), b.get("devise"))),
        ("Prix unitaire EUR", _money(a.get("prix_unitaire_eur")), _money(b.get("prix_unitaire_eur"))),
        ("PayPal G&S", a.get("paypal_gs"), b.get("paypal_gs")),
        ("Tracking", a.get("tracking"), b.get("tracking")),
        ("État", a.get("etat"), b.get("etat")),
        ("Doublons", a.get("doublons"), b.get("doublons")),
        ("Photos / vidéo", a.get("photos_video"), b.get("photos_video")),
        ("Confiance", a.get("confiance_note"), b.get("confiance_note")),
        ("Service", a.get("service_note"), b.get("service_note")),
        ("Contenu", a.get("contenu_note"), b.get("contenu_note")),
        ("Statut", a.get("statut"), b.get("statut")),
        ("Spécialités", ", ".join(a.get("specialites", [])), ", ".join(b.get("specialites", []))),
        ("Position marché", a.get("market_position") or "Non renseignée", b.get("market_position") or "Non renseignée"),
        ("Risque", a.get("risk_level", "unknown"), b.get("risk_level", "unknown")),
        ("Première commande", a.get("recommended_first_order") or "Non renseignée", b.get("recommended_first_order") or "Non renseignée"),
        ("Score PokéStock", f"{scores[a['id']]['score']:.0f}/100" if scores[a["id"]]["score"] is not None else "Score incomplet", f"{scores[b['id']]['score']:.0f}/100" if scores[b["id"]]["score"] is not None else "Score incomplet"),
        ("Couverture", f"{scores[a['id']]['coverage']*100:.0f}%", f"{scores[b['id']]['coverage']*100:.0f}%"),
    ]
    st.table(_prepare_supplier_display_rows([{"Critère": r[0], "A": r[1], "B": r[2]} for r in rows]))
    if a.get("prix_unitaire_eur") is not None and b.get("prix_unitaire_eur") is not None:
        winner = a if a["prix_unitaire_eur"] < b["prix_unitaire_eur"] else b
        st.success(f"Avantage prix selon les données renseignées : {winner.get('nom')}")
    else:
        st.warning("Comparaison prix impossible sans conversion EUR complète.")


def _best_price_supplier(suppliers):
    candidates = [
        supplier for supplier in suppliers
        if supplier.get("statut") not in {"archivé", "À éviter"}
        and supplier.get("landed_cost_per_card_eur") is not None
        and supplier.get("landed_cost_per_card_eur") > 0
    ]
    if candidates:
        return min(candidates, key=lambda item: item.get("landed_cost_per_card_eur")), "co?t rendu France"
    return None, "Co?t rendu France n?cessaire"

def _best_global_supplier(suppliers):
    candidates = []
    for supplier in suppliers:
        if supplier.get("statut") in {"archivé", "à éviter"}:
            continue
        score = pokestock_score(supplier, suppliers)
        if score.get("complete") and score.get("coverage", 0) >= 0.70 and score.get("score") is not None:
            candidates.append((score.get("score"), supplier, score))
    if not candidates:
        return None, None
    _score, supplier, score = max(candidates, key=lambda item: item[0])
    return supplier, score


def _render_supplier_kpi(label, value, hint="", tone="neutral"):
    st.markdown(
        f"""
        <div class="supplier-kpi supplier-kpi-{tone}">
            <span>{html.escape(str(label))}</span>
            <strong>{html.escape(str(value))}</strong>
            <small>{html.escape(str(hint or ""))}</small>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _render_top_two_rankings(suppliers):
    best_price, best_price_hint = _best_price_supplier(suppliers)
    best_global, best_global_score = _best_global_supplier(suppliers)
    st.markdown("### Meilleurs fournisseurs")
    c1, c2 = st.columns(2)
    if best_price:
        c1.markdown(
            f"""
            <div class="supplier-ranking-card supplier-ranking-price">
                <span class="supplier-ranking-icon">🏷️</span>
                <small>Meilleur prix</small>
                <h3>{html.escape(best_price.get('nom') or 'Sans nom')}</h3>
                <strong>{html.escape(_landed_cost_label(best_price))}</strong>
                <p>{html.escape(best_price.get('nom') or '')} · {html.escape(best_price_hint)}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        c1.markdown(
            """
            <div class="supplier-ranking-card supplier-ranking-price">
                <span class="supplier-ranking-icon">🏷️</span>
                <small>Meilleur prix</small>
                <h3>Conversion EUR nécessaire</h3>
                <p>Les lignes conservent les prix source quand ils existent.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    if best_global:
        conf = confidence_score(best_global)
        c2.markdown(
            f"""
            <div class="supplier-ranking-card supplier-ranking-global">
                <span class="supplier-ranking-icon">🏆</span>
                <small>Meilleur choix global</small>
                <h3>{html.escape(best_global.get('nom') or 'Sans nom')}</h3>
                <strong>Score PokéStock : {best_global_score.get('score'):.0f} / 100</strong>
                <p>Risque {_risk_label(best_global.get('risk_level')).lower()} · Confiance {conf:.1f} / 5</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        c2.markdown(
            """
            <div class="supplier-ranking-card supplier-ranking-global">
                <span class="supplier-ranking-icon">🏆</span>
                <small>Meilleur choix global</small>
                <h3>Pas encore assez de données.</h3>
                <p>Le score doit être complet avec au moins 70 % de couverture.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )


def _best_secure_supplier(suppliers):
    candidates = [
        item for item in suppliers
        if item.get("landed_cost_per_card_eur") is not None
        and item.get("risk_level") in {"low", "medium"}
        and item.get("paypal_gs") == "oui"
        and item.get("tracking") == "oui"
        and item.get("statut") not in {"archivé", "à éviter"}
    ]
    return min(candidates, key=lambda item: item.get("landed_cost_per_card_eur")) if candidates else None


def _best_negotiation_supplier(suppliers):
    candidates = [
        item for item in suppliers
        if (item.get("potentiel_negociation_note") or 0) >= 4
        and item.get("risk_level") in {"low", "medium", "unknown"}
        and item.get("statut") not in {"archivé", "à éviter"}
    ]
    return max(candidates, key=lambda item: ((item.get("potentiel_negociation_note") or 0), _priority_score(item))) if candidates else None


def _render_decision_tile(title, supplier, subtitle=""):
    with st.container(border=True):
        st.caption(title)
        if not supplier:
            st.markdown("**Données insuffisantes**")
            st.caption(subtitle or "Aucun fournisseur ne remplit encore les critères.")
            return
        st.markdown(f"**{supplier.get('nom') or 'Sans nom'}**")
        st.caption(_landed_cost_label(supplier))
        st.caption(subtitle or _display_action_short(supplier.get("action_recommandee")))


def _supplier_action_bucket(supplier):
    decision = _display_action_short(supplier.get("action_recommandee"))
    if supplier.get("statut") == "à éviter" or "pause" in decision.lower() or "éviter" in decision.lower():
        return "À éviter ou mettre en pause"
    if supplier.get("follow_status") in {"En attente de réponse", "En pause"}:
        return "À surveiller"
    if supplier.get("follow_up_date") or supplier.get("follow_status") in {"À contacter", "Négociation", "Devis reçu"}:
        return "À contacter maintenant"
    if "demander" in decision.lower() or "continuer" in decision.lower() or "relancer" in decision.lower():
        return "À contacter maintenant"
    return "À surveiller"


def _render_supplier_action_row(data, supplier, settings, key_prefix):
    metrics = _scenario_metrics(supplier, settings)
    c1, c2, c3, c4 = st.columns([2.4, 1.4, 2.4, 2])
    c1.markdown(f"**{supplier.get('nom') or 'Sans nom'}**")
    c1.caption(_landed_cost_label(supplier))
    c2.caption("Risque")
    c2.markdown(_risk_badge(supplier.get("risk_level")), unsafe_allow_html=True)
    c3.caption(_display_action_short(supplier.get("action_recommandee")))
    c3.caption(f"Marge scénario : {_money_eur(metrics.get('margin')) if metrics.get('margin') is not None else 'Incomplète'}")
    if c4.button("Marquer comme fait", key=f"{key_prefix}_done_{supplier['id']}"):
        supplier = dict(supplier)
        supplier["follow_status"] = "En pause"
        supplier["follow_up_note"] = "Action marquée comme faite"
        supplier["last_interaction_at"] = now_iso()
        data = upsert_supplier(data, supplier, add_history=False, source_type="manual")
        st.success("Action marquée comme faite.")
        st.rerun()
    new_date = c4.text_input("Reporter", value=supplier.get("follow_up_date") or "", key=f"{key_prefix}_date_{supplier['id']}", placeholder="YYYY-MM-DD")
    if c4.button("Ajouter relance", key=f"{key_prefix}_follow_{supplier['id']}"):
        supplier = dict(supplier)
        supplier["follow_up_date"] = new_date
        supplier["follow_status"] = "En attente de réponse"
        data = upsert_supplier(data, supplier, add_history=False, source_type="manual")
        st.success("Relance enregistrée.")
        st.rerun()


def _render_test_order_sheet(supplier, settings):
    metrics = _scenario_metrics(supplier, settings)
    with st.expander("Préparer une commande test", expanded=False):
        st.caption("Fiche de suivi uniquement : aucune commande réelle, aucun Stock/Lot/Vente modifié.")
        rows = [
            {"Champ": "Fournisseur", "Valeur": supplier.get("nom") or "Sans nom"},
            {"Champ": "Variante", "Valeur": supplier.get("offer_variant") or "Non renseignée"},
            {"Champ": "Quantité", "Valeur": supplier.get("quantity_reference") or 200},
            {"Champ": "Coût rendu France estimé", "Valeur": _landed_total_label(supplier)},
            {"Champ": "Revente estimée", "Valeur": _money_eur(metrics.get("resale"))},
            {"Champ": "Marge estimée", "Valeur": _money_eur(metrics.get("margin")) if metrics.get("margin") is not None else "Incomplète"},
            {"Champ": "% achat / cote", "Valeur": f"{metrics.get('purchase_pct'):.1f} %" if metrics.get("purchase_pct") is not None else "Incomplet"},
        ]
        st.table(_prepare_supplier_display_rows(rows))
        checklist = supplier.get("test_order_checklist") or []
        for item in checklist:
            st.checkbox(item.get("label", "Point à vérifier"), value=bool(item.get("done")), disabled=True, key=f"test_order_{supplier['id']}_{item.get('label')}")


def _render_reviews_overview(data):
    history = data.get("review_history", [])
    if not history:
        st.info("Aucune review importée pour le moment.")
        return
    latest = history[0]
    with st.container(border=True):
        st.caption("Dernière review importée")
        st.markdown(f"**{latest.get('supplier_name') or 'Fournisseur'}**")
        st.caption(f"{latest.get('format')} · {latest.get('result')} · {latest.get('created_at')}")
        st.caption(latest.get("summary") or "Résumé indisponible")
        with st.expander("Voir la review", expanded=False):
            st.text_area("Texte brut", latest.get("source_text", ""), height=160, disabled=True, key=f"latest_review_{latest.get('id')}")
    with st.expander("Dernières reviews", expanded=False):
        for item in history[:5]:
            st.markdown(f"**{item.get('supplier_name') or 'Fournisseur'}** · {item.get('result')} · {item.get('created_at')}")
            st.caption(item.get("summary") or "Résumé indisponible")
            with st.expander("Texte brut", expanded=False):
                st.text_area("Source", item.get("source_text", ""), height=120, disabled=True, key=f"review_raw_{item.get('id')}")


def render_fournisseurs_page(context):
    globals().update(context)
    st.markdown(
        render_page_header("Fournisseurs", "Comparer vos sources AR / CHR", "🤝"),
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <style>
        .supplier-meta {color:#64748b;font-weight:700;font-size:.82rem;margin:.2rem 0 .55rem}
        .supplier-import-card {background:linear-gradient(135deg,#ffffff,#f5f3ff);border:1px solid #ddd6fe;border-radius:18px;padding:1rem;box-shadow:0 16px 36px rgba(76,29,149,.10);margin:.8rem 0 1rem}
        .supplier-success-card {background:#ecfdf5;border:1px solid #86efac;border-radius:16px;padding:.9rem 1rem;margin:.75rem 0;color:#14532d}
        .supplier-success-card h4 {margin:0 0 .35rem;font-size:1rem;color:#166534}
        .supplier-kpi {border-radius:16px;padding:.9rem 1rem;border:1px solid #e2e8f0;background:#fff;box-shadow:0 10px 26px rgba(15,23,42,.06);min-height:112px}
        .supplier-kpi span {display:block;color:#64748b;font-size:.72rem;font-weight:900;text-transform:uppercase;letter-spacing:.04em}
        .supplier-kpi strong {display:block;color:#111827;font-size:1.35rem;margin:.25rem 0}
        .supplier-kpi small {color:#64748b;font-weight:700}
        .supplier-kpi-price {border-color:#67e8f9;background:linear-gradient(135deg,#ecfeff,#fff)}
        .supplier-kpi-score {border-color:#c4b5fd;background:linear-gradient(135deg,#f5f3ff,#fff)}
        .supplier-ranking-card {position:relative;overflow:hidden;border-radius:18px;padding:1.1rem;border:1px solid #e2e8f0;box-shadow:0 14px 34px rgba(15,23,42,.08);min-height:174px}
        .supplier-ranking-card small {font-weight:900;text-transform:uppercase;letter-spacing:.05em;color:#64748b}
        .supplier-ranking-card h3 {margin:.35rem 0;color:#111827}
        .supplier-ranking-card strong {font-size:1.05rem;color:#0f172a}
        .supplier-ranking-card p {margin:.45rem 0 0;color:#475569;font-weight:700}
        .supplier-ranking-icon {font-size:1.6rem;float:right}
        .supplier-ranking-price {background:linear-gradient(135deg,#ecfeff,#ffffff);border-color:#67e8f9}
        .supplier-ranking-global {background:linear-gradient(135deg,#f5f3ff,#ffffff);border-color:#c4b5fd}
        .supplier-calc {display:grid;gap:.2rem;background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:.7rem;margin:.45rem 0;font-size:.86rem}
        .supplier-calc-total {font-weight:900;color:#4c1d95}
        @media(max-width:760px){.supplier-table-desktop{display:none}}
        </style>
        """,
        unsafe_allow_html=True,
    )

    data = load_suppliers(create_if_missing=True)
    data, repair_summary = repair_existing_supplier_reviews(data)
    if not repair_summary.get("already_done") and repair_summary.get("suppliers_completed"):
        st.success(f"Reviews réanalysées : {repair_summary.get('suppliers_completed')} fournisseurs complétés.")
    suppliers = data.get("suppliers", [])

    with st.expander("Taux de conversion local", expanded=False):
        settings = data.setdefault("settings", {})
        rates = settings.setdefault("conversion_rates", {})
        jpy = rates.setdefault("JPY", {"rate_to_eur": 0.0062, "updated_at": ""})
        usd = rates.setdefault("USD", {"rate_to_eur": 0.92, "updated_at": ""})
        r1, r2, r3 = st.columns(3)
        jpy_rate = r1.number_input("JPY → EUR", min_value=0.0, value=float(jpy.get("rate_to_eur") or 0.0062), step=0.0001, format="%.6f")
        usd_rate = r2.number_input("USD → EUR", min_value=0.0, value=float(usd.get("rate_to_eur") or 0.92), step=0.01, format="%.4f")
        rate_date = r3.text_input("Dernière mise à jour du taux", value=jpy.get("updated_at") or usd.get("updated_at") or "")
        scenario = st.selectbox(
            "Scénario de revente",
            options=["prudent", "reference", "optimiste", "personnalise"],
            format_func=lambda value: {
                "prudent": "Prudent · 2,50 €",
                "reference": "Référence · 3,00 €",
                "optimiste": "Optimiste · 3,50 €",
                "personnalise": "Personnalisé",
            }.get(value, value),
            index=["prudent", "reference", "optimiste", "personnalise"].index(settings.get("resale_scenario", "reference")) if settings.get("resale_scenario", "reference") in ["prudent", "reference", "optimiste", "personnalise"] else 1,
            key="supplier_resale_scenario",
        )
        custom_resale = st.number_input("Revente personnalisée / carte", min_value=0.0, value=float(settings.get("custom_resale_per_card_eur") or 3.0), step=0.10)
        if st.button("Enregistrer les taux locaux", key="save_supplier_rates"):
            settings["conversion_rates"]["JPY"] = {"rate_to_eur": jpy_rate, "updated_at": rate_date}
            settings["conversion_rates"]["USD"] = {"rate_to_eur": usd_rate, "updated_at": rate_date}
            settings["resale_scenario"] = scenario or "reference"
            settings["custom_resale_per_card_eur"] = custom_resale
            data = save_suppliers(data)
            st.success("Taux locaux enregistrés. Aucun taux n'a été récupéré sur internet.")
            st.rerun()

    st.markdown("### Importer une review fournisseur")
    with st.container(border=True):
        quick_key_version = st.session_state.get("supplier_quick_import_key_version", 0)
        quick_key = f"supplier_quick_import_text_{quick_key_version}"
        quick_text = st.text_area(
            "Colle une review, un update ou un devis",
            height=150,
            key=quick_key,
            placeholder="=== SUPPLIER REVIEW V1 ===\nSupplier: Japan TCG Exchange\nOffer: 200 cards: ¥44,800\nRisk Level: Low",
        )
        st.caption("Analyse locale uniquement. Ne collez pas de mots de passe, coordonnées bancaires ou informations sensibles.")
        q1, q2 = st.columns([2, 1])
        if q1.button("Importer automatiquement", type="primary", width="stretch"):
            try:
                data, result = apply_supplier_import(data, quick_text, automatic=True)
                st.session_state["supplier_last_import_result"] = result
                if result.get("status") in {"created", "updated", "family_updated"}:
                    st.session_state.pop("supplier_import_preview", None)
                    st.session_state.pop("supplier_import_text", None)
                    st.session_state["supplier_quick_import_key_version"] = quick_key_version + 1
                st.rerun()
            except Exception as exc:
                st.session_state["supplier_last_import_result"] = {"status": "error", "message": str(exc)}
        if q2.button("Voir avant import", width="stretch"):
            st.session_state["supplier_import_preview"] = extract_offer_from_text(quick_text)
            st.session_state["supplier_import_text"] = quick_text
        last_result = st.session_state.get("supplier_last_import_result")
        if last_result:
            status = last_result.get("status")
            parsed = last_result.get("parsed", {})
            supplier = last_result.get("supplier") or parsed.get("supplier") or {}
            if status in {"created", "updated"}:
                action = "créé" if status == "created" else "mis à jour"
                st.markdown(
                    f"""
                    <div class="supplier-success-card">
                        <h4>{html.escape(supplier.get('nom') or 'Fournisseur')} {action} ✓</h4>
                        <p>Prix 200 : {html.escape(_money(supplier.get('prix_200'), supplier.get('devise') or 'EUR'))}</p>
                        <p>Risque : {_risk_label(supplier.get('risk_level'))} · Confiance : {html.escape(_clean_text(supplier.get('confiance_note')))} / 5</p>
                        <p>Action : {html.escape(_display_action(supplier.get('action_recommandee')))}</p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                status = "__handled_success"
            if status == "created":
                st.success(f"Fournisseur créé ✓ {supplier.get('nom') or ''}")
            elif status == "updated":
                st.success(f"Fournisseur mis à jour ✓ {supplier.get('nom') or ''}")
            elif status == "family_updated":
                st.success(
                    f"{last_result.get('supplier_family_name') or 'Famille fournisseur'} mis à jour ✓ "
                    f"Mise à jour générale appliquée à {last_result.get('linked_count', 0)} offre(s) liée(s)."
                )
            elif status == "duplicate":
                st.info("Review déjà importée pour ce fournisseur.")
            elif status == "pending":
                st.warning("Import placé dans Imports à vérifier : " + (last_result.get("reason") or "vérification nécessaire"))
            elif status == "duplicate_other":
                st.warning("Review déjà appliquée à un autre fournisseur : vérification recommandée.")
            elif status == "error":
                st.error("Import impossible : " + (last_result.get("message") or "erreur inconnue"))
    suppliers = data.get("suppliers", [])

    active = [s for s in suppliers if s.get("statut") != "archivé"]
    tested = [s for s in suppliers if s.get("statut") == "testé"]
    avoid = [s for s in suppliers if s.get("statut") == "à éviter"]
    best_price_supplier, best_price_hint = _best_price_supplier(suppliers)
    best_global_supplier, best_global_score = _best_global_supplier(suppliers)
    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        _render_supplier_kpi("Actifs", len(active), "Fournisseurs disponibles", "neutral")
    with k2:
        _render_supplier_kpi(
            "Meilleur prix",
            _supplier_unit_price_label(best_price_supplier) if best_price_supplier else "—",
            f"{best_price_supplier.get('nom')} · {best_price_hint}" if best_price_supplier else best_price_hint,
            "price",
        )
    with k3:
        _render_supplier_kpi(
            "Meilleur score",
            f"{best_global_score.get('score'):.0f} / 100" if best_global_score else "—",
            best_global_supplier.get("nom") if best_global_supplier else "Score incomplet",
            "score",
        )
    with k4:
        _render_supplier_kpi("Testés", len(tested), "Statut confirmé", "neutral")
    with k5:
        _render_supplier_kpi("À éviter", len(avoid), "Statut manuel", "neutral")

    pending_imports = data.get("pending_imports", [])
    if pending_imports:
        with st.expander(f"Imports à vérifier · {len(pending_imports)}", expanded=True):
            for pending in pending_imports:
                st.markdown(f"**{pending.get('supplier_name') or 'Nom non détecté'}** · {pending.get('format')} · {pending.get('reason')}")
                summary = pending.get("summary", {})
                st.caption(
                    f"Prix 200 : {_money(summary.get('prix_200'), summary.get('devise') or 'EUR')} · "
                    f"Dernière offre : {_money(summary.get('latest_offer'), summary.get('devise') or 'EUR')} · "
                    f"Risque : {summary.get('risk_level') or 'unknown'}"
                )
                target_options = ["Choisir un fournisseur"] + [
                    f"{supplier.get('nom') or 'Sans nom'} · {supplier.get('id')}"
                    for supplier in suppliers
                ]
                target_choice = st.selectbox(
                    "Appliquer à un fournisseur existant",
                    target_options,
                    key=f"pending_target_{pending['id']}",
                )
                p1, p2, p3, p4 = st.columns(4)
                if p1.button("Créer un nouveau fournisseur", key=f"pending_create_{pending['id']}"):
                    candidate = pending.get("parsed_supplier") or default_supplier()
                    candidate.setdefault("import_history", []).append({
                        "id": f"import_{pending.get('import_hash', '')[:12]}",
                        "date": now_iso(),
                        "import_hash": pending.get("import_hash"),
                        "source_type": pending.get("format"),
                        "source_text": pending.get("source_text", ""),
                    })
                    data = upsert_supplier(data, candidate, add_history=True, source_type=pending.get("format") or "message_import", source_text=pending.get("source_text", ""))
                    data = delete_pending_import(data, pending["id"])
                    st.success("Import ambigu créé après confirmation.")
                    st.rerun()
                if p2.button("Appliquer", key=f"pending_apply_{pending['id']}"):
                    if target_choice == "Choisir un fournisseur":
                        st.warning("Choisis d'abord le fournisseur cible.")
                    else:
                        target_id = target_choice.rsplit(" · ", 1)[-1]
                        existing = next((item for item in suppliers if item.get("id") == target_id), None)
                        if existing:
                            candidate = pending.get("parsed_supplier") or {}
                            merged = dict(existing)
                            for field in pending.get("detected_fields", []):
                                if field in {"id", "created_at", "updated_at", "notes"}:
                                    continue
                                value = candidate.get(field)
                                if value in (None, "", [], "unknown"):
                                    continue
                                merged[field] = value
                            merged.setdefault("import_history", list(existing.get("import_history", []))).append({
                                "id": f"import_{pending.get('import_hash', '')[:12]}",
                                "date": now_iso(),
                                "import_hash": pending.get("import_hash"),
                                "source_type": pending.get("format"),
                                "source_text": pending.get("source_text", ""),
                            })
                            data = upsert_supplier(
                                data,
                                normalize_supplier(merged),
                                add_history=True,
                                source_type=pending.get("format") or "message_import",
                                source_text=pending.get("source_text", ""),
                            )
                            data = delete_pending_import(data, pending["id"])
                            st.success("Import appliqué après confirmation.")
                            st.rerun()
                if p3.button("Modifier les champs", key=f"pending_edit_{pending['id']}"):
                    st.session_state["supplier_import_text"] = pending.get("source_text", "")
                    st.session_state["supplier_import_preview"] = extract_offer_from_text(pending.get("source_text", ""))
                    st.info("Aperçu prêt dans l'onglet Importer depuis message / review.")
                if p4.button("Supprimer l'import", key=f"pending_delete_{pending['id']}"):
                    data = delete_pending_import(data, pending["id"])
                    st.rerun()

    tab_overview, tab_list, tab_actions, tab_directory = st.tabs(["Vue d’ensemble", "Comparer", "À traiter", "Répertoire"])

    with tab_overview:
        _render_reviews_overview(data)
        o1, o2, o3, o4 = st.columns(4)
        best_landed, _best_hint = _best_price_supplier(suppliers)
        best_secure = _best_secure_supplier(suppliers)
        best_global, _best_score = _best_global_supplier(suppliers)
        best_nego = _best_negotiation_supplier(suppliers)
        with o1:
            _render_decision_tile("Meilleur prix rendu France", best_landed)
        with o2:
            _render_decision_tile("Meilleur prix sécurisé", best_secure, "Risque faible/moyen ? PayPal G&S ? tracking")
        with o3:
            _render_decision_tile("Meilleur choix global", best_global, "Score complet ou priorité recommand?e")
        with o4:
            _render_decision_tile("Opportunit? ? négocier", best_nego, "Potentiel élevé et risque acceptable")
        buckets = {"À contacter maintenant": [], "À surveiller": [], "À éviter ou mettre en pause": []}
        for item in suppliers:
            if item.get("statut") == "archivé":
                continue
            buckets.setdefault(_supplier_action_bucket(item), []).append(item)
        for title, items in buckets.items():
            with st.expander(f"{title} ? {len(items)}", expanded=(title == "À contacter maintenant")):
                for item in sorted(items, key=lambda s: _priority_score(s), reverse=True)[:8]:
                    _render_supplier_action_row(data, item, data.get("settings", {}), f"overview_{title}")

    with tab_directory:
        editing_id = st.selectbox(
            "Fournisseur à modifier",
            ["Nouveau"] + [f"{s.get('nom') or 'Sans nom'} · {s.get('id')}" for s in suppliers],
            key="supplier_edit_select",
        )
        selected = None
        if editing_id != "Nouveau":
            selected_id = editing_id.rsplit(" · ", 1)[-1]
            selected = next((s for s in suppliers if s.get("id") == selected_id), None)
        result = _supplier_form("supplier_edit", selected, "Enregistrer le fournisseur")
        if result:
            if not result.get("nom"):
                st.error("Nom fournisseur requis.")
            else:
                data = upsert_supplier(data, result, add_history=bool(result.pop("_save_history", False)), source_type="manual")
                st.success("Fournisseur enregistré.")
                st.rerun()
        if selected:
            c1, c2 = st.columns(2)
            if c1.button("Dupliquer l'offre dans l'historique", width="stretch"):
                data = duplicate_offer(data, selected["id"])
                st.success("Offre ajoutée à l'historique.")
                st.rerun()
            confirm_key = f"confirm_delete_supplier_{selected['id']}"
            if not st.session_state.get(confirm_key):
                if c2.button("Supprimer", width="stretch"):
                    st.session_state[confirm_key] = True
                    st.rerun()
            else:
                st.warning("Confirmer la suppression de ce fournisseur ?")
                d1, d2 = st.columns(2)
                if d1.button("Oui, supprimer", type="primary", width="stretch"):
                    data = delete_supplier(data, selected["id"])
                    st.session_state.pop(confirm_key, None)
                    st.success("Fournisseur supprimé.")
                    st.rerun()
                if d2.button("Annuler", width="stretch"):
                    st.session_state.pop(confirm_key, None)
                    st.rerun()

    with tab_directory:
        st.markdown("### Importer depuis message / review")
        source_text = st.text_area("Message, devis, conversation ou review", height=180, key="supplier_import_text")
        st.caption("Analyse locale uniquement. Ne collez pas de mots de passe, coordonnées bancaires ou informations sensibles.")
        if st.button("Analyser et pré-remplir", type="primary"):
            st.session_state["supplier_import_preview"] = extract_offer_from_text(source_text)
        preview = st.session_state.get("supplier_import_preview")
        if preview:
            candidate = preview["supplier"]
            detected_fields = preview.get("detected_fields") or []
            duplicate = detect_duplicate(data, candidate)
            near_matches = supplier_name_matches(data.get("suppliers", []), candidate.get("nom", ""))
            if duplicate and all(item.get("id") != duplicate.get("id") for item in near_matches):
                near_matches.insert(0, duplicate)
            st.info("Analyse locale uniquement : aucune donnée n'a été envoyée vers un service externe.")
            st.success(_import_format_label(preview.get("import_format")))
            if detected_fields:
                st.markdown("**Champs détectés**")
                st.table(_prepare_supplier_display_rows([{"Champ": _field_label(field), "Valeur": _display_value(candidate.get(field))} for field in detected_fields if field != "analysis_updated_at"]))
            expected_fields = {"nom", "pays", "type", "statut", "prix_100", "prix_200", "devise", "port", "paiement", "paypal_gs", "frais_paypal", "tracking", "photos_video", "etat", "doublons", "confiance_note", "service_note", "contenu_note", "potentiel_negociation_note", "derniere_analyse", "action_recommandee", "market_position", "best_for", "risk_level", "recommended_first_order", "review_last_updated"}
            absent_fields = sorted(expected_fields - set(detected_fields))
            if absent_fields:
                st.caption("Champs absents : " + ", ".join(_field_label(field) for field in absent_fields[:12]) + ("..." if len(absent_fields) > 12 else ""))
            if preview.get("uncertain_fields"):
                st.warning("Champs incertains : " + ", ".join(preview["uncertain_fields"]))
            if preview.get("warnings"):
                for warning in preview["warnings"]:
                    st.warning(warning)
            if len(near_matches) > 1:
                st.warning("Plusieurs fournisseurs possibles détectés : choisis le bon avant toute mise à jour.")
                match_label = st.selectbox(
                    "Fournisseur existant concerné",
                    [f"{item.get('nom') or 'Sans nom'} · {item.get('id')}" for item in near_matches],
                    key="supplier_import_match_select",
                )
                match_id = match_label.rsplit(" · ", 1)[-1]
                duplicate = next((item for item in near_matches if item.get("id") == match_id), None)
            if duplicate:
                st.warning(f"Fournisseur existant possible : {duplicate.get('nom')}")
                differences = _candidate_differences(duplicate, candidate, detected_fields)
                if differences:
                    st.markdown("**Différences détectées**")
                    st.table(_prepare_supplier_display_rows(differences))
                else:
                    st.caption("Aucune différence nette détectée sur les champs reconnus.")
            imported = _supplier_form("supplier_import", candidate, "Créer / enregistrer depuis l'import")
            if imported:
                if not imported.get("nom"):
                    st.error("Nom fournisseur requis avant sauvegarde.")
                elif duplicate:
                    imported["_source_text"] = source_text
                    imported["_duplicate_id"] = duplicate["id"]
                    imported["_detected_fields"] = detected_fields
                    imported["_import_format"] = preview.get("import_format") or "message_import"
                    st.session_state["supplier_import_pending_save"] = imported
                    st.info("Fournisseur existant détecté : choisis l'action à appliquer avant d'enregistrer.")
                else:
                    add_history = bool(imported.pop("_save_history", False))
                    if add_history:
                        imported = _with_price_history(
                            imported,
                            make_price_history_entry(None, imported, source_type=preview.get("import_format") or "message_import", source_text=source_text),
                        )
                    data = upsert_supplier(data, imported, add_history=add_history, source_type="message_import", source_text=source_text)
                    st.session_state.pop("supplier_import_preview", None)
                    st.success("Import enregistré après confirmation.")
                    st.rerun()
            pending = st.session_state.get("supplier_import_pending_save")
            if pending:
                st.markdown("**Action sur le fournisseur existant**")
                a1, a2, a3, a4 = st.columns(4)
                source_text_pending = pending.get("_source_text", source_text)
                duplicate_id = pending.get("_duplicate_id")
                save_history_flag = bool(pending.get("_save_history", False))
                detected_fields_pending = pending.get("_detected_fields") or []
                import_format_pending = pending.get("_import_format") or "message_import"
                if a1.button("Créer un nouveau fournisseur", width="stretch"):
                    payload = {k: v for k, v in pending.items() if not str(k).startswith("_")}
                    if save_history_flag:
                        payload = _with_price_history(
                            payload,
                            make_price_history_entry(None, payload, source_type=import_format_pending, source_text=source_text_pending),
                        )
                    data = upsert_supplier(data, payload, add_history=save_history_flag, source_type="message_import", source_text=source_text_pending)
                    st.session_state.pop("supplier_import_pending_save", None)
                    st.session_state.pop("supplier_import_preview", None)
                    st.success("Nouveau fournisseur créé.")
                    st.rerun()
                if a2.button("Mettre à jour l'existant", width="stretch"):
                    payload = {k: v for k, v in pending.items() if not str(k).startswith("_")}
                    existing = next((s for s in data.get("suppliers", []) if s.get("id") == duplicate_id), None)
                    if existing:
                        payload = _merge_detected_fields(existing, payload, detected_fields_pending)
                        payload = _with_price_history(
                            payload,
                            make_price_history_entry(existing, payload, source_type=import_format_pending, source_text=source_text_pending),
                        )
                    data = upsert_supplier(data, payload, add_history=True, source_type="message_import", source_text=source_text_pending)
                    st.session_state.pop("supplier_import_pending_save", None)
                    st.session_state.pop("supplier_import_preview", None)
                    st.success("Fournisseur existant mis à jour avec une nouvelle entrée d'historique.")
                    st.rerun()
                if a3.button("Ajouter offre à l'historique", width="stretch"):
                    payload = {k: v for k, v in pending.items() if not str(k).startswith("_")}
                    existing = next((s for s in data.get("suppliers", []) if s.get("id") == duplicate_id), None)
                    if existing:
                        payload = dict(existing)
                        payload = _with_price_history(
                            payload,
                            make_price_history_entry(existing, pending, source_type=import_format_pending, source_text=source_text_pending),
                        )
                    data = upsert_supplier(data, payload, add_history=True, source_type="message_import", source_text=source_text_pending)
                    st.session_state.pop("supplier_import_pending_save", None)
                    st.session_state.pop("supplier_import_preview", None)
                    st.success("Offre ajoutée à l'historique du fournisseur existant.")
                    st.rerun()
                if a4.button("Créer une nouvelle offre", width="stretch"):
                    payload = {k: v for k, v in pending.items() if not str(k).startswith("_")}
                    existing = next((s for s in data.get("suppliers", []) if s.get("id") == duplicate_id), None)
                    if existing:
                        payload = _merge_detected_fields(existing, payload, detected_fields_pending)
                        payload = _with_price_history(
                            payload,
                            make_price_history_entry(existing, payload, source_type=import_format_pending, source_text=source_text_pending),
                        )
                    data = upsert_supplier(data, payload, add_history=True, source_type="message_import", source_text=source_text_pending)
                    st.session_state.pop("supplier_import_pending_save", None)
                    st.session_state.pop("supplier_import_preview", None)
                    st.success("Nouvelle offre enregistrée après confirmation.")
                    st.rerun()
        if st.button("Annuler l'import"):
            st.session_state.pop("supplier_import_preview", None)
            st.session_state.pop("supplier_import_pending_save", None)
            st.rerun()

    with tab_actions:
        st.markdown("### À traiter")
        action_suppliers = [s for s in suppliers if s.get("statut") != "archivé"]
        for item in sorted(action_suppliers, key=lambda s: _priority_score(s), reverse=True)[:20]:
            _render_supplier_action_row(data, item, data.get("settings", {}), "actions")

    with tab_list:
        if not suppliers:
            st.info("Aucun fournisseur pour le moment. Ajoute un fournisseur ou importe un message pour commencer.")
            return
        sort_mode = st.selectbox(
            "Tri",
            ["Priorité recommandée", "Prix unitaire EUR croissant", "Confiance décroissante", "Service décroissant", "Contenu décroissant", "Position marché", "Statut", "Pays", "Date de mise à jour"],
            key="supplier_sort",
        )
        search_query = st.text_input("Recherche fournisseur", key="supplier_search_query", placeholder="Nom, pays, statut...")
        country_options = sorted({s.get("pays") for s in suppliers if s.get("pays")})
        active_filter_count = sum([
            bool(st.session_state.get("supplier_search_query")),
            st.session_state.get("supplier_country_filter", "Tous") != "Tous",
            st.session_state.get("supplier_type_filter", "Tous") != "Tous",
            st.session_state.get("supplier_status_filter", "Tous") != "Tous",
            st.session_state.get("supplier_currency_filter", "Toutes") != "Toutes",
            st.session_state.get("supplier_paypal_filter", "Tous") != "Tous",
            st.session_state.get("supplier_media_filter", "Tous") != "Tous",
            st.session_state.get("supplier_tracking_filter", "Tous") != "Tous",
            st.session_state.get("supplier_min_trust_filter", 0.0) > 0,
            st.session_state.get("supplier_risk_filter", "Tous les risques") != "Tous les risques",
            st.session_state.get("supplier_first_order_filter", "Toutes") != "Toutes",
            st.session_state.get("supplier_max_price_filter", 0.0) > 0,
        ])
        with st.expander(f"Filtres avancés · {active_filter_count} actif(s)", expanded=False):
            c1, c2, c3, c4 = st.columns(4)
            country_filter = c1.selectbox("Pays", ["Tous"] + country_options, key="supplier_country_filter")
            type_filter = c2.selectbox("Type", ["Tous"] + SUPPLIER_TYPES, key="supplier_type_filter")
            status_filter = c3.selectbox("Statut", ["Tous"] + SUPPLIER_STATUSES, key="supplier_status_filter")
            currency_filter = c4.selectbox("Devise", ["Toutes"] + SUPPLIER_CURRENCIES, key="supplier_currency_filter")
            f1, f2, f3, f4 = st.columns(4)
            paypal_filter = f1.selectbox("PayPal G&S", ["Tous"] + PAYPAL_STATES, key="supplier_paypal_filter")
            media_filter = f2.selectbox("Photos / vidéo", ["Tous"] + MEDIA_STATES, key="supplier_media_filter")
            tracking_filter = f3.selectbox("Tracking", ["Tous"] + TRACKING_STATES, key="supplier_tracking_filter")
            min_trust_filter = f4.slider("Confiance min.", 0.0, 5.0, 0.0, 0.5, key="supplier_min_trust_filter")
            r1, r2 = st.columns(2)
            risk_filter = r1.selectbox("Niveau de risque", ["Tous les risques", "Faible", "Moyen", "Élevé", "Non renseigné"], key="supplier_risk_filter")
            first_order_filter = r2.selectbox(
                "Première commande recommandée",
                ["Toutes", "100 cartes ou moins", "200 cartes", "Plus de 200 cartes", "Non renseigné"],
                key="supplier_first_order_filter",
            )
            max_price_filter = st.number_input("Prix unitaire EUR max. (0 = aucun filtre)", min_value=0.0, value=0.0, step=0.05, key="supplier_max_price_filter")
            if st.button("Réinitialiser les filtres", key="supplier_reset_filters"):
                for key in [
                    "supplier_search_query", "supplier_country_filter", "supplier_type_filter", "supplier_status_filter",
                    "supplier_currency_filter", "supplier_paypal_filter", "supplier_media_filter", "supplier_tracking_filter",
                    "supplier_min_trust_filter", "supplier_risk_filter", "supplier_first_order_filter", "supplier_max_price_filter",
                ]:
                    st.session_state.pop(key, None)
                st.rerun()
        filtered = suppliers
        if search_query:
            query = search_query.casefold()
            filtered = [
                s for s in filtered
                if query in " ".join([
                    str(s.get("nom") or ""),
                    str(s.get("pays") or ""),
                    str(s.get("type") or ""),
                    str(s.get("statut") or ""),
                    str(s.get("action_recommandee") or ""),
                ]).casefold()
            ]
        if country_filter != "Tous":
            filtered = [s for s in filtered if s.get("pays") == country_filter]
        if type_filter != "Tous":
            filtered = [s for s in filtered if s.get("type") == type_filter]
        if status_filter != "Tous":
            filtered = [s for s in filtered if s.get("statut") == status_filter]
        if currency_filter != "Toutes":
            filtered = [s for s in filtered if s.get("devise") == currency_filter]
        if paypal_filter != "Tous":
            filtered = [s for s in filtered if s.get("paypal_gs") == paypal_filter]
        if media_filter != "Tous":
            filtered = [s for s in filtered if s.get("photos_video") == media_filter]
        if tracking_filter != "Tous":
            filtered = [s for s in filtered if s.get("tracking") == tracking_filter]
        risk_map = {"Faible": "low", "Moyen": "medium", "Élevé": "high", "Non renseigné": "unknown"}
        if risk_filter != "Tous les risques":
            filtered = [s for s in filtered if s.get("risk_level", "unknown") == risk_map[risk_filter]]
        if first_order_filter == "100 cartes ou moins":
            filtered = [s for s in filtered if s.get("recommended_first_order_quantity") is not None and s.get("recommended_first_order_quantity") <= 100]
        elif first_order_filter == "200 cartes":
            filtered = [s for s in filtered if s.get("recommended_first_order_quantity") == 200]
        elif first_order_filter == "Plus de 200 cartes":
            filtered = [s for s in filtered if s.get("recommended_first_order_quantity") is not None and s.get("recommended_first_order_quantity") > 200]
        elif first_order_filter == "Non renseigné":
            filtered = [s for s in filtered if s.get("recommended_first_order_quantity") is None]
        if min_trust_filter > 0:
            filtered = [s for s in filtered if (s.get("confiance_note") or 0) >= min_trust_filter]
        if max_price_filter > 0:
            filtered = [s for s in filtered if s.get("prix_unitaire_eur") is not None and s.get("prix_unitaire_eur") <= max_price_filter]
        if sort_mode == "Priorité recommandée":
            filtered = sorted(filtered, key=lambda s: (_priority_score(s), s.get("updated_at") or ""), reverse=True)
        elif sort_mode == "Prix unitaire EUR croissant":
            filtered = sorted(filtered, key=lambda s: s.get("prix_unitaire_eur") if s.get("prix_unitaire_eur") is not None else 999999)
        elif sort_mode == "Confiance décroissante":
            filtered = sorted(filtered, key=lambda s: confidence_score(s) if confidence_score(s) is not None else -1, reverse=True)
        elif sort_mode == "Service décroissant":
            filtered = sorted(filtered, key=lambda s: s.get("service_note") if s.get("service_note") is not None else -1, reverse=True)
        elif sort_mode == "Contenu décroissant":
            filtered = sorted(filtered, key=lambda s: s.get("contenu_note") if s.get("contenu_note") is not None else -1, reverse=True)
        elif sort_mode == "Position marché":
            filtered = sorted(filtered, key=lambda s: (s.get("market_position_rank") is None, s.get("market_position_rank") or 999999, s.get("market_position") or ""))
        elif sort_mode == "Statut":
            filtered = sorted(filtered, key=lambda s: s.get("statut") or "")
        elif sort_mode == "Pays":
            filtered = sorted(filtered, key=lambda s: s.get("pays") or "")
        else:
            filtered = sorted(filtered, key=lambda s: s.get("updated_at") or "", reverse=True)

        duplicates = duplicate_supplier_groups(data)
        if duplicates:
            with st.expander(f"Doublons potentiels détectés · {len(duplicates)}", expanded=False):
                st.caption("Aucune fusion automatique : choisis explicitement la fiche à conserver.")
                for group_index, group in enumerate(duplicates):
                    identity = group.get("identity") or f"group_{group_index}"
                    options = group.get("suppliers", [])
                    names = [f"{item.get('nom') or 'Sans nom'} :: {item.get('id')}" for item in options]
                    st.markdown("**" + " / ".join(item.get("nom") or "Sans nom" for item in options) + "**")
                    base_ids = "|".join(item.get("id", "") for item in options)
                    keep_key = _supplier_widget_key("merge_keep", group_index, identity, base_ids)
                    merge_key = _supplier_widget_key("merge_source", group_index, identity, base_ids)
                    keep_label = st.selectbox("Fournisseur ? conserver", names, key=keep_key)
                    keep_id = keep_label.rsplit(" :: ", 1)[-1]
                    merge_names = [name for name in names if not name.endswith(f":: {keep_id}")]
                    merge_label = st.selectbox("Fournisseur ? fusionner", merge_names, key=merge_key)
                    merge_id = merge_label.rsplit(" :: ", 1)[-1]
                    st.caption("Résumé : les historiques, reviews, aliases et champs manquants seront conservés ; la fiche absorbée sera archivée.")
                    confirm_key = _supplier_widget_key("merge_confirm", group_index, identity, keep_id, merge_id)
                    button_key = _supplier_widget_key("merge_apply", group_index, identity, keep_id, merge_id)
                    ignore_key = _supplier_widget_key("merge_ignore", group_index, identity, keep_id, merge_id)
                    confirm = st.checkbox("Je confirme la fusion manuelle de ce doublon", key=confirm_key)
                    if st.button("Comparer puis fusionner", key=button_key, disabled=not confirm or keep_id == merge_id):
                        data = merge_suppliers(data, keep_id, merge_id)
                        st.success("Doublon fusionné après confirmation. La fiche absorbée est archivée.")
                        st.rerun()
                    if st.button("Pas un doublon", key=ignore_key, disabled=keep_id == merge_id):
                        data = ignore_duplicate_pair(data, keep_id, merge_id)
                        st.success("Paire ignorée pour les prochaines détections.")
                        st.rerun()

        st.markdown("### Comparatif fournisseurs")
        comparison_rows = _comparison_rows(filtered)
        if comparison_rows:
            primary_columns = ["Fournisseur", "Coût rendu / carte", "% achat / cote", "Marge estimée", "Prix cartes €", "Livraison €", "PayPal €", "Total fournisseur €", "Risque", "Priorité", "Décision"]
            st.dataframe(
                _prepare_supplier_display_rows([{key: row.get(key) for key in primary_columns} for row in comparison_rows]),
                width="stretch",
                hide_index=True,
            )
            st.download_button(
                "Exporter le comparatif Excel",
                data=_build_suppliers_workbook(filtered),
                file_name="comparatif_fournisseurs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )
        else:
            st.info("Aucun fournisseur ne correspond aux filtres.")

        with st.expander("Importer un tableau fournisseurs Excel", expanded=False):
            uploaded = st.file_uploader("Fichier .xlsx", type=["xlsx"], key="supplier_excel_import")
            if uploaded:
                parsed_excel = _parse_excel_suppliers(uploaded)
                st.caption(f"{len(parsed_excel)} fournisseur(s) détecté(s). Aperçu avant import obligatoire.")
                preview_rows = [
                    {
                        "Fournisseur": item["supplier"].get("nom") or "Sans nom",
                        "Pays": _clean_text(item["supplier"].get("pays"), ""),
                        "Prix 200": _money(item["supplier"].get("prix_200"), item["supplier"].get("devise") or "EUR"),
                        "Devise": item["supplier"].get("devise") or "",
                        "Risque": _risk_label(item["supplier"].get("risk_level")),
                        "Action recommandée": _display_action(item["supplier"].get("action_recommandee")),
                    }
                    for item in parsed_excel[:20]
                ]
                st.dataframe(_prepare_supplier_display_rows(preview_rows), width="stretch", hide_index=True)
                if st.button("Importer le tableau confirmé", type="primary"):
                    created = updated_count = pending_count = 0
                    for item in parsed_excel:
                        supplier = item["supplier"]
                        matches = supplier_name_matches(data.get("suppliers", []), supplier.get("nom", ""))
                        if len(matches) > 1:
                            data.setdefault("pending_imports", []).append({
                                "id": f"pending_excel_{supplier.get('id')}",
                                "created_at": now_iso(),
                                "import_hash": "",
                                "format": "excel_import",
                                "reason": "fournisseur Excel ambigu",
                                "supplier_name": supplier.get("nom"),
                                "summary": {"prix_200": supplier.get("prix_200"), "devise": supplier.get("devise"), "risk_level": supplier.get("risk_level")},
                                "source_text": "",
                                "parsed_supplier": supplier,
                                "detected_fields": item.get("detected_fields", []),
                                "warnings": [],
                            })
                            pending_count += 1
                        elif len(matches) == 1:
                            supplier["id"] = matches[0].get("id")
                            supplier["created_at"] = matches[0].get("created_at")
                            data = upsert_supplier(data, _merge_detected_fields(matches[0], supplier, item.get("detected_fields", [])), add_history=False, source_type="excel_import")
                            updated_count += 1
                        else:
                            data = upsert_supplier(data, supplier, add_history=False, source_type="excel_import")
                            created += 1
                    if pending_count:
                        data = save_suppliers(data)
                    st.success(f"Import Excel terminé : {created} créé(s), {updated_count} mis à jour, {pending_count} à vérifier.")
                    st.rerun()

        _render_top_two_rankings(filtered)
        with st.expander("Classements spécialisés et comparaison avancée", expanded=False):
            _render_rankings(filtered)
            _render_smart_rankings(filtered)
            st.markdown("### Comparer deux fournisseurs")
            _render_compare(filtered)
        with st.expander(f"Fiches fournisseurs · {len(filtered)}", expanded=False):
            for supplier in filtered:
                _render_supplier_summary_card(supplier, suppliers)
                with st.expander(f"Détail · {supplier.get('nom') or 'Sans nom'}", expanded=False):
                    overview_tab, neg_tab, conv_tab, hist_tab, score_tab = st.tabs(
                        ["Aperçu", "Négociations", "Conversation", "Historique des offres", "Score & classements"]
                    )
                    with overview_tab:
                        d1, d2 = st.columns(2)
                        with d1:
                            st.markdown("**Calcul 200 cartes**")
                            _render_calculation(supplier, 200)
                        with d2:
                            st.markdown("**Calcul 100 cartes**")
                            _render_calculation(supplier, 100)
                        st.caption(f"Contact : {supplier.get('contact_link') or 'Non renseigné'}")
                        st.caption(f"État : {supplier.get('etat')} · Doublons : {supplier.get('doublons')} · Tracking : {supplier.get('tracking')}")
                        st.caption("Spécialités : " + (", ".join(supplier.get("specialites") or []) or "Non renseignées"))
                        potential = supplier.get("potentiel_negociation_note")
                        st.caption(f"Potentiel de négociation : {potential:.1f}/5" if potential is not None else "Potentiel de négociation : non renseigné")
                        if supplier.get("derniere_analyse") or supplier.get("action_recommandee"):
                            st.markdown("**Dernière analyse**")
                            st.caption(_clean_text(supplier.get("derniere_analyse"), "Aucune analyse enregistrée."))
                            st.markdown("**Action recommandée**")
                            st.caption(_display_action(supplier.get("action_recommandee")) or "Aucune action recommandée.")
                            if supplier.get("analysis_updated_at"):
                                st.caption(f"Dernière mise à jour : {supplier.get('analysis_updated_at')}")
                        else:
                            st.caption("Aucune analyse enregistrée.")
                        if supplier.get("pros"):
                            st.caption("Points forts : " + ", ".join(supplier.get("pros")))
                        if supplier.get("cons"):
                            st.caption("Points faibles : " + ", ".join(supplier.get("cons")))
                        st.markdown("**Métadonnées review**")
                        meta_a, meta_b, meta_c = st.columns(3)
                        rank = supplier.get("market_position_rank")
                        market_position = supplier.get("market_position") or "Non renseignée"
                        meta_a.caption("Position marché")
                        meta_a.caption(f"{market_position}" + (f" · #{int(rank)}" if rank is not None else ""))
                        meta_b.caption("Idéal pour")
                        meta_b.caption(_clean_text(supplier.get("best_for"), "Non renseigné"))
                        if supplier.get("best_for_tags"):
                            meta_b.markdown(" ".join(_badge(tag, "info") for tag in supplier.get("best_for_tags")), unsafe_allow_html=True)
                        meta_c.caption("Niveau de risque")
                        meta_c.markdown(_risk_badge(supplier.get("risk_level")), unsafe_allow_html=True)
                        order_text = supplier.get("recommended_first_order") or "Non renseignée"
                        order_qty = supplier.get("recommended_first_order_quantity")
                        st.caption(
                            "Première commande recommandée : "
                            + (f"{int(order_qty)} cartes · {order_text}" if order_qty else order_text)
                        )
                        review_date = _format_review_date(supplier.get("review_last_updated"))
                        if review_date:
                            st.caption(f"Review mise à jour le : {review_date}")
                        _render_score(supplier, suppliers)
                        if supplier.get("notes"):
                            st.text_area("Notes", supplier.get("notes"), disabled=True, key=f"notes_view_{supplier['id']}")
                    with neg_tab:
                        _render_negotiations(data, supplier)
                    with conv_tab:
                        _render_conversation(data, supplier)
                    with hist_tab:
                        if supplier.get("price_history"):
                            st.markdown("**Historique prix**")
                            for entry in reversed(supplier["price_history"]):
                                old_price = _money(entry.get("ancien_prix"), entry.get("devise")) if entry.get("ancien_prix") is not None else "non comparable"
                                new_price = _money(entry.get("nouveau_prix"), entry.get("devise"))
                                st.caption(
                                    f"{entry.get('date')} · {entry.get('source_type')} · "
                                    f"{old_price} → {new_price} · quantité {entry.get('quantite') or '—'}"
                                )
                                if entry.get("analyse"):
                                    st.caption("Analyse : " + entry.get("analyse"))
                                if entry.get("action_recommandee"):
                                    st.caption("Action : " + entry.get("action_recommandee"))
                        if supplier.get("offers_history"):
                            st.markdown("**Historique des offres**")
                            for entry in reversed(supplier["offers_history"]):
                                st.caption(
                                    f"{entry.get('date')} · {entry.get('source_type')} · "
                                    f"200: {_money(entry.get('prix_final_estime'), entry.get('devise'))} · "
                                    f"unitaire EUR: {_money(entry.get('prix_unitaire_eur'))}"
                                )
                        else:
                            st.caption("Aucune offre historisée.")
                    with score_tab:
                        st.markdown("**Score PokéStock**")
                        st.caption("Classement indicatif basé sur vos prix, vos notes et votre historique.")
                        _render_score(supplier, suppliers)
                        stats = negotiation_stats(supplier)
                        if stats.get("comparable_count"):
                            st.caption(f"Potentiel de négociation : {stats['comparable_count']} négociation(s), meilleure baisse {stats.get('best_reduction_pct'):.1f} %")
                        else:
                            st.caption("Potentiel de négociation : historique insuffisant.")
